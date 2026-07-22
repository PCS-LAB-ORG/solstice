"""
dashboard.py — Solstice Control Center
Dark ops dashboard inspired by Pi-hole + OpenClaw aesthetic.
FastAPI + SSE + Chart.js. No npm required.
Run: python3 dashboard.py → http://localhost:8200

Security note: innerHTML is used only for server-generated HTML from
trusted SQLite data, never from user input. All user-facing dynamic
values are inserted via textContent.
"""

import json, asyncio, subprocess, sys
from datetime import datetime, timezone, date, timedelta
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

from contextlib import asynccontextmanager
from fastapi import FastAPI, HTTPException, Request
from fastapi.responses import HTMLResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
import uvicorn

from agent.constants import STATE_FILE, DATA_DIR
from agent.db import get_db, init_db


@asynccontextmanager
async def lifespan(app):
    """Populate DB immediately when FastAPI starts."""
    loop = asyncio.get_event_loop()
    await loop.run_in_executor(None, _populate_db)
    yield


app = FastAPI(title="Solstice Control Center", lifespan=lifespan)
app.mount(
    "/static",
    StaticFiles(directory=str(Path(__file__).parent / "static")),
    name="static",
)


def _populate_db():
    """Populate DB from CSVs. Checks all critical tables, not just blocked_data."""
    try:
        with get_db() as conn:
            bd = conn.execute("SELECT COUNT(*) FROM blocked_data").fetchone()[0]
            sh = conn.execute("SELECT COUNT(*) FROM status_history").fetchone()[0]
            ai = conn.execute("SELECT COUNT(*) FROM ai_enrichment").fetchone()[0]
        if bd > 0 and sh > 0:
            return  # All tables populated (ai_enrichment intentionally unused since Ollama removed)
    except:
        pass

    import json as _j, logging as _log

    _log.warning("DB empty — populating from CSVs...")
    from agent.ps_parser import load_and_merge as _mp
    from agent.db import migrate_from_state as _mig

    init_db()
    if (DATA_DIR / "ps_tracker.csv").exists():
        _mp(DATA_DIR / "ps_tracker.csv", STATE_FILE)
    _mig(STATE_FILE)
    # DC pipeline runs last — syncs all 9 milestones, cc_rep, cc_dsm, churn_risk, rebuilds m1_suggestions
    if (DATA_DIR / "unified_tracker2.xlsx").exists():
        _run_dc_pipeline(DATA_DIR, STATE_FILE)
    _state = _j.loads(STATE_FILE.read_text())
    with get_db() as conn:
        for aid, acc in _state.get("accounts", {}).items():
            conn.execute(
                "UPDATE accounts SET live_fire=?, live_fire_dc=? WHERE account_id=?",
                (
                    1 if acc.get("live_fire") else 0,
                    acc.get("live_fire_dc", "") or "",
                    aid,
                ),
            )
    with get_db() as conn:
        n2 = conn.execute("SELECT COUNT(*) FROM blocked_data").fetchone()[0]
    _log.warning("DB populated: %d milestone records", n2)


def _ensure_db():
    """Call before every API response."""
    _populate_db()


def _cohort_sql(alias: str = "b") -> str:
    """Returns SQL fragment for cohort filter. Bind with (cohort_val, cohort_val)."""
    return f"AND (? = '' OR {alias}.cohort = ?)"


def _load_milestones_sla_count() -> list:
    """Fast SLA breach count used by stats."""
    try:
        return [
            r
            for r in _load_milestones()
            if r.get("sla_m3_m8_breach") or r.get("sla_m8_m9_breach")
        ]
    except:
        return []


def _load_stats() -> dict:
    _ensure_db()
    try:
        with get_db() as conn:
            total = conn.execute(
                "SELECT COUNT(*) FROM accounts WHERE customer_name!=''"
            ).fetchone()[0]
            by_status = {
                r[0] or "Blank": r[1]
                for r in conn.execute(
                    "SELECT status, COUNT(*) FROM accounts WHERE customer_name!='' GROUP BY status"
                )
            }
            # DC milestones are source of truth for completion/progress
            completed = conn.execute(
                "SELECT COUNT(*) FROM blocked_data WHERE m9_complete=1"
            ).fetchone()[0]
            in_prog_dc = conn.execute(
                "SELECT COUNT(*) FROM blocked_data WHERE m8_started=1 AND (m9_complete IS NULL OR m9_complete=0)"
            ).fetchone()[0]
            no_status = sum(v for k, v in by_status.items() if k in ("", "Blank"))
            stalls = conn.execute("""SELECT COUNT(DISTINCT b.account_id)
                FROM blocked_data b JOIN accounts a ON a.account_id=b.account_id
                WHERE b.is_cs_team=1 AND b.m3_complete=1 AND b.m8_started=0""").fetchone()[
                0
            ]
            core_rep = conn.execute(
                "SELECT COUNT(*) FROM blocked_data WHERE subtype='core_rep_blocking'"
            ).fetchone()[0]
            ps_active = conn.execute(
                "SELECT COUNT(*) FROM ps_data WHERE ps_status!='' AND ps_status NOT LIKE 'BLOCKED%'"
            ).fetchone()[0]
            no_cse = conn.execute(
                "SELECT COUNT(*) FROM accounts WHERE (active_cse IS NULL OR active_cse='') AND customer_name!=''"
            ).fetchone()[0]
            dc_driving = conn.execute(
                "SELECT COUNT(*) FROM accounts WHERE live_fire=1"
            ).fetchone()[0]
            sla_breaches = len(_load_milestones_sla_count())
            dc_green = conn.execute(
                "SELECT COUNT(*) FROM blocked_data WHERE dc_progress='Green'"
            ).fetchone()[0]
            dc_yellow = conn.execute(
                "SELECT COUNT(*) FROM blocked_data WHERE dc_progress='Yellow'"
            ).fetchone()[0]
            dc_red = conn.execute(
                "SELECT COUNT(*) FROM blocked_data WHERE dc_progress='Red'"
            ).fetchone()[0]
            last_run = STATE_FILE.stat().st_mtime if STATE_FILE.exists() else 0
            last_str = (
                datetime.fromtimestamp(last_run, timezone.utc).strftime(
                    "%d %b %Y · %H:%M UTC"
                )
                if last_run
                else "—"
            )
            return dict(
                total=total,
                completed=completed,
                in_progress=in_prog_dc,
                no_status=no_status,
                stalls=stalls,
                core_rep_blocking=core_rep,
                ps_active=ps_active,
                no_cse=no_cse,
                dc_driving=dc_driving,
                by_status=by_status,
                last_run=last_str,
                sla_breaches=sla_breaches,
                dc_green=dc_green,
                dc_yellow=dc_yellow,
                dc_red=dc_red,
            )
    except Exception as e:
        return {"error": str(e)}


def _load_weekly() -> list:
    _ensure_db()
    try:
        today = date.today()
        weeks = {}
        with get_db() as conn:
            for row in conn.execute("""SELECT a.customer_name,a.active_cse,a.status,
                a.live_fire, a.live_fire_dc, a.sales_region, a.account_theatre,
                b.m8_planned,b.m9_planned,b.m8_started,b.m9_complete
                FROM accounts a JOIN blocked_data b ON a.account_id=b.account_id
                WHERE b.m3_complete=1
                AND (b.m8_started=1 OR b.m8_planned!='' OR b.m9_planned!='')"""):
                for field, label, done in [
                    (row["m8_planned"], "M8", bool(row["m8_started"])),
                    (
                        row["m9_planned"],
                        "M9",
                        bool(row["m9_complete"]) or row["status"] == "Completed",
                    ),
                ]:
                    if not field:
                        continue
                    try:
                        for fmt in ("%m/%d/%Y", "%m/%d/%y"):
                            try:
                                d = datetime.strptime(field.strip(), fmt).date()
                                break
                            except:
                                pass
                        else:
                            continue
                        k = d.strftime("%Y-W%U")
                        ws = d - timedelta(days=(d.weekday() + 1) % 7)
                        weeks.setdefault(
                            k,
                            {
                                "key": k,
                                "label": f"W{d.strftime('%U')} · {ws.strftime('%d %b')}",
                                "m8": [],
                                "m8_done": [],
                                "m9": [],
                                "done": [],
                            },
                        )
                        e = {
                            "name": row["customer_name"],
                            "cse": row["active_cse"] or "—",
                            "lf": bool(row["live_fire"]),
                            "lf_dc": row["live_fire_dc"] or "",
                            "date": d.strftime("%d %b"),
                            "region": row["sales_region"] or "",
                            "theatre": row["account_theatre"] or "EMEA",
                        }
                        if label == "M8" and not done:
                            weeks[k]["m8"].append(e)
                        elif label == "M8" and done:
                            weeks[k]["m8_done"].append(e)
                        elif label == "M9" and not done:
                            weeks[k]["m9"].append(e)
                        elif label == "M9" and done:
                            weeks[k]["done"].append(e)
                    except:
                        pass
        tw = today.strftime("%Y-W%U")
        sw = sorted(weeks.values(), key=lambda x: x["key"])
        idx = next((i for i, w in enumerate(sw) if w["key"] >= tw), 0)
        return sw[max(0, idx - 3) : idx + 6]
    except:
        return []


def _load_cse() -> list:
    _ensure_db()
    try:
        with get_db() as conn:
            rows = conn.execute("""SELECT active_cse,COUNT(*) total,
                SUM(CASE WHEN status IN ('Ready To Engage','Account team contacted') THEN 1 ELSE 0 END) outreach,
                SUM(CASE WHEN status IN ('Sales Hold','Churning/Churned','Backoff','Cancelled') THEN 1 ELSE 0 END) risk,
                SUM(CASE WHEN status='Completed' THEN 1 ELSE 0 END) done,
                SUM(CASE WHEN status IN ('In Progress','Kick Off Scheduled','Blocked: Tech limitation',
                    'Blocked: Acct Team','Blocked: Customer','Customer Engaged') THEN 1 ELSE 0 END) active
                FROM accounts WHERE active_cse IS NOT NULL AND active_cse!=''
                GROUP BY active_cse ORDER BY total DESC""").fetchall()
            result = []
            for r in rows:
                d = dict(r)
                eligible = d["total"] - d["risk"]
                d["success_rate"] = (
                    round(d["done"] / eligible * 100) if eligible > 0 else 0
                )
                result.append(d)
            return result
    except:
        return []


@app.get("/api/stats")
def api_stats():
    return _load_stats()


@app.get("/api/weekly")
def api_weekly():
    return _load_weekly()


@app.get("/api/cse")
def api_cse():
    return _load_cse()


_XSUP_CLOSED = {"Closed", "Done", "Engineering Resolved"}


def _parse_and_store_xsup(xlsx_path: Path) -> int:
    """
    Parse Open XSUPs tab from xlsx, drop closed statuses, upsert into xsup_data.
    Returns count of open XSUPs stored.
    Runs account_id matching against accounts table by name (fuzzy-tolerant lower strip).
    """
    try:
        import openpyxl as _xl
    except ImportError:
        return 0

    wb = _xl.load_workbook(str(xlsx_path), read_only=True, data_only=True)
    ws = wb["Open XSUPs"]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return 0

    headers = [str(h).strip() if h is not None else "" for h in rows[0][:22]]
    data = []
    for row in rows[1:]:
        rec = dict(zip(headers, row[:22]))
        if not any(v for v in rec.values() if v is not None and str(v).strip()):
            continue
        status = str(rec.get("XSUP Status") or "").strip()
        if status in _XSUP_CLOSED or not status:
            continue
        acct = str(rec.get("Account Name") or "").strip()
        if not acct:
            continue
        data.append(rec)

    now = datetime.now(timezone.utc).isoformat()
    with get_db() as conn:
        # Build name→account_id lookup (lower-stripped)
        acct_map = {
            str(r["customer_name"]).strip().lower(): r["account_id"]
            for r in conn.execute("SELECT account_id, customer_name FROM accounts")
            if r["customer_name"]
        }
        conn.execute("DELETE FROM xsup_data")
        for rec in data:
            acct_name = str(rec.get("Account Name") or "").strip()
            account_id = acct_map.get(acct_name.lower())
            conn.execute(
                """
                INSERT INTO xsup_data
                  (account_name, account_id, case_number, case_status, case_theatre,
                   xsup_number, xsup_priority, xsup_status, summary, component, notes, synced_at)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
                """,
                (
                    acct_name,
                    account_id,
                    str(rec.get("Case Number") or "").strip(),
                    str(rec.get("Case Status") or "").strip(),
                    str(rec.get("Case Theatre") or "").strip(),
                    str(rec.get("XSUP Number") or "").strip(),
                    str(rec.get("XSUP Priority") or "").strip(),
                    str(rec.get("XSUP Status") or "").strip(),
                    str(rec.get("Summary") or "").strip()[:400],
                    str(rec.get("Component") or "").strip(),
                    str(rec.get("Notes") or "").strip()[:600],
                    now,
                ),
            )
    return len(data)


def _parse_and_store_coe(xlsx_bytes: bytes) -> tuple[int, int]:
    """
    Parse Sheet1 (feature/blocker requests) and Cortex Bugs from the Central
    Technical COE Tracker xlsx. Drops and repopulates coe_issues + coe_bugs.
    Returns (issue_count, bug_count).
    """
    try:
        import openpyxl as _xl, io as _io
    except ImportError:
        return (0, 0)

    wb = _xl.load_workbook(_io.BytesIO(xlsx_bytes), data_only=True)
    now = datetime.now(timezone.utc).isoformat()

    def _s(v):
        return str(v or "").strip()

    # ── Sheet1 ────────────────────────────────────────────────────────────────
    ws1 = wb["Sheet1"]
    rows1 = list(ws1.iter_rows(values_only=True))
    h1 = [_s(c) for c in rows1[0]]

    def _c1(row, name):
        try:
            return _s(row[h1.index(name)])
        except (ValueError, IndexError):
            return ""

    issue_rows = []
    for row in rows1[1:]:
        if not any(v for v in row if v is not None and str(v).strip()):
            continue
        issue_rows.append(
            (
                _c1(row, "Issue ID"),
                _c1(row, "Timestamp"),
                _c1(row, "Upgrade Blocker"),
                _c1(row, "Request Type"),
                _c1(row, "DC Assigned (TRR)"),
                _c1(row, "Core Theater"),
                _c1(row, "Core Area"),
                _c1(row, "Account Name"),
                _c1(row, "Technical issue")[:600],
                _c1(row, "Requirements")[:600],
                _c1(row, "Priority"),
                _c1(row, "Module"),
                _c1(row, "Issue Category"),
                _c1(row, "Resource Name"),
                _c1(row, "Issue Notes (PM/Engg/SPO DC)")[:600],
                _c1(row, "Status - Has the question been answered definitively?"),
                _c1(row, "Timeline/ Answer")[:400],
                _c1(row, "Outcome - is this in the product currently?"),
                _c1(row, "Top 100"),
                now,
            )
        )

    # ── Cortex Bugs ───────────────────────────────────────────────────────────
    ws2 = wb["Cortex Bugs"]
    rows2 = list(ws2.iter_rows(values_only=True))

    bug_rows = []
    for row in rows2[1:]:
        if not any(v for v in row if v is not None and str(v).strip()):
            continue
        bug_rows.append(
            (
                _s(row[0]),  # account_name
                _s(row[1]),  # xsup_number
                _s(row[2]),  # xsup_assignee
                _s(row[3]),  # xsup_priority
                _s(row[4]),  # xsup_status
                _s(row[5]),  # spo_dc_classification
                _s(row[6]),  # eng_escalation_status
                _s(row[7]),  # component
                _s(row[8])[:400],  # summary
                _s(row[9])[:600],  # notes
                now,
            )
        )

    import re as _re

    def _norm(s):
        s = s.lower().strip()
        s = _re.sub(
            r"\b(limited|ltd|inc|corp|corporation|llc|plc|sa|spa|nv|bv|ab|as|oy|gmbh|kg|s\.a\.|s\.p\.a\.)\b\.?",
            "",
            s,
        )
        s = _re.sub(r"[^a-z0-9\s]", " ", s)
        return " ".join(s.split())

    with get_db() as conn:
        # Build account_id lookup (exact + normalised)
        _acct_exact = {
            str(r["customer_name"]).strip().lower(): r["account_id"]
            for r in conn.execute(
                "SELECT account_id, customer_name FROM accounts WHERE customer_name IS NOT NULL"
            )
        }
        _acct_norm = {_norm(k): v for k, v in _acct_exact.items()}

        def _match_account(name):
            if not name:
                return None
            if name.lower() in _acct_exact:
                return _acct_exact[name.lower()]
            n = _norm(name)
            if n in _acct_norm:
                return _acct_norm[n]
            for k, v in _acct_norm.items():
                if n[:20] and k.startswith(n[:20]):
                    return v
            return None

        conn.execute("DELETE FROM coe_issues")
        conn.executemany(
            """INSERT INTO coe_issues
               (issue_id,timestamp,upgrade_blocker,request_type,dc_assigned,theatre,area,
                account_name,account_id,technical_issue,requirements,priority,module,issue_category,
                resource_name,issue_notes,status,timeline_answer,outcome,top_100,synced_at)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            [r[:8] + (_match_account(r[7]),) + r[8:] for r in issue_rows],
        )
        conn.execute("DELETE FROM coe_bugs")
        conn.executemany(
            """INSERT INTO coe_bugs
               (account_name,account_id,xsup_number,xsup_assignee,xsup_priority,xsup_status,
                spo_dc_classification,eng_escalation_status,component,summary,notes,synced_at)
               VALUES (?,?,?,?,?,?,?,?,?,?,?,?)""",
            [r[:1] + (_match_account(r[0]),) + r[1:] for r in bug_rows],
        )

    return (len(issue_rows), len(bug_rows))


def _gdrive_root() -> Path:
    """Resolve Google Drive root — works on Mac (native) and in Docker (volume mount)."""
    docker_mount = Path("/root/gdrive")
    if docker_mount.exists():
        return docker_mount
    return Path.home() / "Library/CloudStorage/GoogleDrive-mbanica@paloaltonetworks.com"


def _download_live_from_drive() -> dict:
    """
    Download Unified Tracker 2.0 xlsx, XSUP Tracker, and COE Tracker from Google Drive
    using ADC token. Reads file IDs from locally synced .gsheet files — no browser needed.
    """
    import json as _j, warnings as _w, subprocess as _sp

    _w.filterwarnings("ignore")
    try:
        import requests as _req
    except ImportError:
        return {"Unified Tracker": "⚠️ requests not installed"}

    result = {}

    # ── Get ADC token (reused for all downloads) ──────────────────────
    try:
        token = _sp.check_output(
            ["gcloud", "auth", "application-default", "print-access-token"],
            text=True,
            stderr=_sp.DEVNULL,
            timeout=10,
        ).strip()
    except Exception as _te:
        result["Unified Tracker"] = f"⚠️ ADC token failed: {_te}"
        return result

    if not token:
        result["Unified Tracker"] = "⚠️ ADC token empty"
        return result

    def _dl_xlsx(url: str, dest: Path) -> str:
        """Download xlsx via ADC token — works in Docker and on host."""
        try:
            r = _req.get(
                url,
                headers={"Authorization": f"Bearer {token}"},
                timeout=120,
                verify=False,
            )
            if r.status_code != 200:
                return f"⚠️ Drive returned {r.status_code}"
            dest.write_bytes(r.content)
            return "ok"
        except Exception as e:
            return f"⚠️ download failed: {e}"

    # ── Unified Tracker 2.0 (replaces DC CSE Tracker CSV) ──────────────
    UNIFIED_GSHEET = (
        _gdrive_root()
        / "My Drive/Cortex Cloud Work"
        / "Unified Tracker 2.0 (IT Connected Sheet).gsheet"
    )
    try:
        unified_file_id = _j.loads(UNIFIED_GSHEET.read_text())["doc_id"]
        unified_dest = DATA_DIR / "unified_tracker2.xlsx"
        dl = _dl_xlsx(
            f"https://docs.google.com/spreadsheets/d/{unified_file_id}/export?format=xlsx",
            unified_dest,
        )
        if dl == "ok":
            result["Unified Tracker"] = f"✅ {unified_dest.stat().st_size // 1024}KB downloaded"
        else:
            result["Unified Tracker"] = dl
    except FileNotFoundError:
        unified_dest = DATA_DIR / "unified_tracker2.xlsx"
        if unified_dest.exists():
            result["Unified Tracker"] = f"✅ loaded from disk ({unified_dest.stat().st_size // 1024}KB)"
        else:
            result["Unified Tracker"] = "⚠️ unified_tracker2.xlsx not found — run host_sync.py"
    except Exception as ue:
        result["Unified Tracker"] = f"⚠️ Unified Tracker download failed: {ue}"

    # Download XSUP tracker
    XSUP_GSHEET = (
        _gdrive_root()
        / "My Drive/Cortex Cloud Work/Cortex Cloud Open XSUPs with TAC.gsheet"
    )
    try:
        xsup_file_id = _j.loads(XSUP_GSHEET.read_text())["doc_id"]
        xsup_dest = DATA_DIR / "xsup_tracker.xlsx"
        dl = _dl_xlsx(
            f"https://docs.google.com/spreadsheets/d/{xsup_file_id}/export?format=xlsx",
            xsup_dest,
        )
        if dl == "ok":
            xsup_rows = _parse_and_store_xsup(xsup_dest)
            result["XSUP Tracker"] = f"✅ {xsup_rows} open XSUPs synced"
        else:
            result["XSUP Tracker"] = dl
    except FileNotFoundError:
        result["XSUP Tracker"] = "⚠️ .gsheet not found — Drive not mounted"
    except Exception as xe:
        result["XSUP Tracker"] = f"⚠️ XSUP download failed: {xe}"

    # Download Central Technical COE Tracker (Sheet1 + Cortex Bugs)
    COE_GSHEET = (
        _gdrive_root()
        / "My Drive/Cortex Cloud Work/Central Technical COE Tracker.gsheet"
    )
    try:
        coe_file_id = _j.loads(COE_GSHEET.read_text())["doc_id"]
        coe_dest = DATA_DIR / "coe_tracker.xlsx"
        dl = _dl_xlsx(
            f"https://docs.google.com/spreadsheets/d/{coe_file_id}/export?format=xlsx",
            coe_dest,
        )
        if dl == "ok":
            coe_counts = _parse_and_store_coe(coe_dest.read_bytes())
            result["COE Tracker"] = (
                f"✅ {coe_counts[0]} issues + {coe_counts[1]} bugs synced"
            )
        else:
            result["COE Tracker"] = dl
    except FileNotFoundError:
        result["COE Tracker"] = "⚠️ .gsheet not found — Drive not mounted"
    except Exception as ce:
        result["COE Tracker"] = f"⚠️ COE download failed: {ce}"

    return result


_DC_MILESTONE_WATCH = [
    ("m0_complete", "M0 Kickoff"),
    ("m1_complete", "M1 Outreach"),
    ("m2_complete", "M2 Entitlements"),
    ("m3_complete", "M3 Buy-in"),
    ("m4_complete", "M4 Discovery"),
    ("m5_complete", "M5 Tech Validation"),
    ("m7_complete", "M7 Legal"),
    ("m8_started", "M8 Upgrade Started"),
    ("m9_complete", "M9 Upgrade Complete"),
]


def _yn_str(v) -> str:
    return (
        "Y"
        if str(v if v is not None else "") in ("1", "True", "true", "Y", "y")
        else "N"
    )


def _run_dc_pipeline(data_dir: Path, state_file: Path) -> dict:
    """
    Unified Tracker 2.0 pipeline step:
    1. Call parse_unified_xlsx — full wipe+replace of accounts + blocked_data
    2. Rebuild m1_suggestions from fresh DB state
    Returns: {matched, total, audit_logged, new_accounts, m1_rebuilt, history_backfilled}
    """
    import json as _j
    from datetime import datetime as _dt, timezone as _tz
    from agent.dc_parser import parse_unified_xlsx as _parse_unified

    # Parse Unified Tracker 2.0 — full wipe+replace of accounts + blocked_data
    unified_path = data_dir / "unified_tracker2.xlsx"
    if not unified_path.exists():
        return {
            "error": "unified_tracker2.xlsx not found — run host_sync.py or Refresh Data",
            "matched": 0,
            "total": 0,
            "audit_logged": 0,
            "new_accounts": 0,
            "m1_rebuilt": 0,
            "history_backfilled": 0,
        }

    with get_db() as _conn:
        _parse_result = _parse_unified(unified_path.read_bytes(), _conn)

    _now = _dt.now(_tz.utc).isoformat()
    _audit = 0
    _snap = {}  # status_history diffing skipped — parse_unified_xlsx handles full wipe+replace

    result = {
        "matched": _parse_result["accounts"],
        "total": _parse_result["accounts"],
    }

    result["audit_logged"] = 0
    result["new_accounts"] = 0

    # Auto-rebuild m1_suggestions from fresh DB state — no stale data ever served
    _SKIP = {"Churning/Churned", "Cancelled", "Backoff", "Completed"}
    _SKIP_CSE = {"PS"}  # PS-owned accounts are not in the scale cohort
    _HOLD_KW = [
        "hold off",
        "please hold",
        "do not reach",
        "strictly on hold",
        "postpone",
        "defer",
        "sales manager",
        "acct team ask",
        "regional sales",
        "advised to hold",
        "core rep is blocking",
    ]
    _REGION_CSE = {
        "UKI": ["Chinmoy Roy", "Tunde Adenugba", "Visnavi"],
        "SEUR": ["Mikhail Bakhmetiev", "Mathieu Dalbes", "Alvaro Fortes", "Visnavi"],
        "France": ["Mathieu Dalbes", "Mikhail Bakhmetiev", "Alvaro Fortes"],
        "Germany": ["Jonathan Brox", "Alvaro Fortes"],
        "Nordics": ["Jonathan Brox", "Mathieu Dalbes"],
        "Benelux": ["Jonathan Brox"],
        "Alps": ["Chinmoy Roy", "Jonathan Brox"],
        "CEE": ["Tunde Adenugba", "Jonathan Brox", "Alvaro Fortes"],
        "Turkey/SA": ["Tunde Adenugba", "Pushkar Kakkar"],
        "Gulf/North Africa": ["Pushkar Kakkar", "Jonathan Brox", "Chinmoy Roy"],
        "Saudi/LBS": ["Jonathan Brox", "Pushkar Kakkar"],
        "—": ["Chinmoy Roy", "Jonathan Brox"],
    }
    with get_db() as _mdb:
        _load = {
            r[0]: r[1]
            for r in _mdb.execute(
                "SELECT active_cse,COUNT(*) FROM accounts WHERE active_cse IS NOT NULL AND active_cse!='' AND active_cse!='Irene Garcia' GROUP BY active_cse"
            ).fetchall()
        }
        _load["Visnavi"] = 0
        _SKIP_KW = [
            "wiz choice",
            "made a wiz",
            "went to wiz",
            "no action required",
            "will not migrate",
            "never activated",
            "not migrate them",
            "decided not to migrate",
            "chose wiz",
            "going with wiz",
            "no cortex cloud migration",
            "nfr use",
            "likely nfr",
            "partner account",
            "strictly nfr",
        ]
        _m1_rows = _mdb.execute("""
            SELECT a.account_id,a.customer_name,a.active_cse,a.sales_region,a.status,a.live_fire,
                   b.signal,b.status_detail,b.upgrade_notes,b.subtype,b.churn_risk,b.health_notes
            FROM accounts a JOIN blocked_data b ON a.account_id=b.account_id
            WHERE a.customer_name!='' AND b.m1_complete=0
              AND b.cohort = 'Scale cohort'
              AND LOWER(a.customer_name) NOT IN (
                SELECT LOWER(a2.customer_name)
                FROM accounts a2 JOIN blocked_data b2 ON a2.account_id=b2.account_id
                WHERE b2.m1_complete=1 AND b2.cohort='Scale cohort'
              )
            ORDER BY a.sales_region,a.customer_name""").fetchall()
        _mdb.execute("DELETE FROM m1_suggestions")
        for _mr in _m1_rows:
            _reg = _mr["sales_region"] or "—"
            _orig = _mr["active_cse"] or "—"
            _cse = (
                min(
                    _REGION_CSE.get(_reg, ["Chinmoy Roy"]),
                    key=lambda c: _load.get(c, 0),
                )
                if _orig == "Irene Garcia"
                else _orig
            )
            if _orig == "Irene Garcia":
                _load[_cse] = _load.get(_cse, 0) + 1
            if _orig in _SKIP_CSE:
                continue
            _st = _mr["status"] or ""
            _sig = _mr["signal"] or ""
            _det = (_mr["status_detail"] or "").lower()
            _nt = (_mr["upgrade_notes"] or "").lower()
            _hn = (_mr["health_notes"] or "").lower()
            _churn = (_mr["churn_risk"] or "").strip()
            _nm = _mr["customer_name"].lower()
            _all = _nt + " " + _hn + " " + _det
            if "mtn " in _nm or _nm == "mtn benin":
                _cat = "unblock"
            elif _st in _SKIP or "Blocked: Tech" in _st:
                _cat = "skip"
            elif any(k in _all for k in _SKIP_KW):
                _cat = "skip"
            elif _churn == "Red":
                _cat = "acct_team"
            elif "\U0001f6d1" in (_mr["status_detail"] or "") or "blocked from" in _det:
                _cat = "acct_team"
            elif (
                _st in ("Sales Hold", "On Hold")
                or any(k in _nt + _det for k in _HOLD_KW)
                or _sig == "blocked"
            ):
                _cat = "acct_team"
            else:
                _cat = "actionable"
            _mdb.execute(
                "INSERT INTO m1_suggestions (account_name,assigned_cse,original_cse,region,status,signal,category) VALUES (?,?,?,?,?,?,?)",
                (
                    _mr["customer_name"],
                    _cse,
                    _orig,
                    _reg,
                    _mr["status"],
                    _mr["signal"],
                    _cat,
                ),
            )
    result["m1_rebuilt"] = len(_m1_rows)

    # History backfill not needed — parse_unified_xlsx does a full wipe+replace each run;
    # status_history table is preserved separately and not wiped.
    result["history_backfilled"] = 0

    # Update last_run in state.json
    _state_data = _j.loads(state_file.read_text())
    _state_data["last_run"] = _dt.now(_tz.utc).isoformat()
    state_file.write_text(_j.dumps(_state_data))

    return result


@app.get("/api/run-pipeline")
def api_run():
    try:
        # Step 1: Download latest from Google Drive via browser
        dl_results = _download_live_from_drive()

        # Step 2: Run pipeline on fresh data (inline, not subprocess — avoids DB connection issues)
        try:
            import json as _j2
            from datetime import datetime as _dt, timezone as _tz
            from agent.ps_parser import load_and_merge as _mp
            from agent.dc_parser import load_and_merge as _mdc
            from agent.enricher import enrich_accounts as _enrich
            from agent.db import migrate_from_state as _mig

            _mp(DATA_DIR / "ps_tracker.csv", STATE_FILE)
            # Unified Tracker 2.0 is master — handles snapshot, upsert, m1_suggestions rebuild
            if (DATA_DIR / "unified_tracker2.xlsx").exists():
                _run_dc_pipeline(DATA_DIR, STATE_FILE)

            _enrich(STATE_FILE)
            _mig(STATE_FILE)
            r_output = "Unified Tracker 2.0 is sole source. All milestones synced.\ndone"
            r_ok = True
        except Exception as _e:
            r_output = str(_e)
            r_ok = False
        return {
            "status": "ok" if r_ok else "error",
            "downloads": dl_results,
            "output": r_output,
        }
    except Exception as e:
        return {"status": "error", "message": str(e)}


@app.get("/api/run-full")
async def api_run_full(request: Request):
    """Stream the full pipeline cycle with debug output."""

    async def stream():
        import os, json as _j, time as _time
        from datetime import datetime, timezone

        _pipeline_start = _time.monotonic()

        def event(step, status, detail="", color="teal"):
            ts = datetime.now().strftime("%H:%M:%S")
            elapsed = _time.monotonic() - _pipeline_start
            return f"data: {_j.dumps({'ts': ts, 'elapsed': round(elapsed, 1), 'step': step, 'status': status, 'detail': detail, 'color': color})}\n\n"

        def step_event(step, status, detail="", color="teal"):
            """Same as event but marks a top-level step for elapsed anchoring."""
            return event(step, status, detail, color)

        yield event("Pipeline", "STARTING", "Full cycle initiated", "teal")
        await asyncio.sleep(0.1)

        # ── Step 1: Data files check (written by host_sync.py into volume) ──
        _t = _time.monotonic()
        data_dir = Path(__file__).parent / "data"
        unified_path = data_dir / "unified_tracker2.xlsx"
        xsup_path = data_dir / "xsup_tracker.xlsx"
        if unified_path.exists():
            import time as _clock

            age_min = int((_clock.time() - unified_path.stat().st_mtime) / 60)
            drive_status = (
                f"unified_tracker2.xlsx {unified_path.stat().st_size // 1024}KB"
                f" · xsup_tracker {'✓' if xsup_path.exists() else '✗'}"
                f" · last updated {age_min}m ago"
            )
            drive_ok = True
        else:
            drive_status = "⚠️ unified_tracker2.xlsx missing — run host_sync.py on Mac host"
            drive_ok = False
        yield event(
            "1/5 Data Files",
            "OK" if drive_ok else "WARN",
            drive_status,
            "blue" if drive_ok else "amber",
        )
        gsheet_files = []
        for gf in gsheet_files:
            try:
                mtime = os.stat(gf).st_mtime
                from datetime import datetime as _dt

                age_h = (_time.time() - mtime) / 3600
                age_str = (
                    f"{age_h:.0f}h ago" if age_h < 48 else f"{age_h / 24:.0f}d ago"
                )
                mt = _dt.fromtimestamp(mtime).strftime("%d %b %H:%M")
                flag = " ⚠️ stale" if age_h > 24 else ""
                yield event(
                    "  →",
                    "FILE",
                    f"{gf.name[:45]} · synced {mt} ({age_str}){flag}",
                    "muted",
                )
            except:
                pass
        yield event(
            "  →", "TIMING", f"Drive check: {_time.monotonic() - _t:.1f}s", "muted"
        )
        await asyncio.sleep(0.1)

        # ── Step 2: Download all sources ──────────────────────────────────
        _t = _time.monotonic()
        data_dir = Path(__file__).parent / "data"

        # Snapshot row counts BEFORE download for delta reporting
        def _snap():
            try:
                with get_db() as _c:
                    return {
                        "xsup": _c.execute("SELECT COUNT(*) FROM xsup_data").fetchone()[
                            0
                        ],
                        "coe_issues": _c.execute(
                            "SELECT COUNT(*) FROM coe_issues"
                        ).fetchone()[0],
                        "coe_bugs": _c.execute(
                            "SELECT COUNT(*) FROM coe_bugs"
                        ).fetchone()[0],
                        "accounts": _c.execute(
                            "SELECT COUNT(*) FROM accounts"
                        ).fetchone()[0],
                    }
            except:
                return {"xsup": 0, "coe_issues": 0, "coe_bugs": 0, "accounts": 0}

        def _delta(before, after):
            d = after - before
            if d > 0:
                return f"+{d} new"
            if d < 0:
                return f"{d} removed"
            return "no change"

        _before = _snap()

        yield event(
            "2/5 Downloading",
            "DOWNLOADING",
            "Pulling Unified Tracker 2.0 · XSUP Tracker · COE Tracker from Drive API",
        )
        dl_results = _download_live_from_drive()
        _after = _snap()

        _dl_ok = 0
        _dl_warn = 0
        for _name, _msg in dl_results.items():
            _ok = "✅" in _msg
            _color = "green" if _ok else "amber"
            if _ok:
                if "XSUP" in _name:
                    _msg += f"  [{_delta(_before['xsup'], _after['xsup'])}]"
                elif "COE" in _name:
                    _msg += f"  [issues: {_delta(_before['coe_issues'], _after['coe_issues'])} · bugs: {_delta(_before['coe_bugs'], _after['coe_bugs'])}]"
                elif "Unified" in _name:
                    _msg += f"  [accounts: {_delta(_before['accounts'], _after['accounts'])}]"
            yield event("  →", "OK" if _ok else "WARN", f"{_name}: {_msg}", _color)
            if _ok:
                _dl_ok += 1
            else:
                _dl_warn += 1
        _dl_elapsed = _time.monotonic() - _t
        _dl_summary = (
            f"{_dl_ok}/{_dl_ok + _dl_warn} sources downloaded · {_dl_elapsed:.1f}s"
        )
        yield event("  →", "TIMING", _dl_summary, "green" if _dl_warn == 0 else "amber")
        await asyncio.sleep(0.2)

        # ── Step 3: Parse Unified Tracker 2.0 ────────────────────────────
        _t = _time.monotonic()
        yield event(
            "3/5 Unified Tracker 2.0",
            "LOADING",
            "Parsing milestones M0–M9 · CSE assignments · all theatres via Unified Tracker 2.0",
        )
        try:
            dc = _run_dc_pipeline(data_dir, data_dir / "state.json")
            _match_pct = round(100 * dc["matched"] / dc["total"]) if dc["total"] else 0
            yield event(
                "  →",
                "OK",
                f"{dc['matched']}/{dc['total']} accounts matched ({_match_pct}%) · "
                f"{dc['audit_logged']} milestone changes · "
                f"M1 rebuilt: {dc.get('m1_rebuilt', 0)} · "
                f"history backfilled: {dc.get('history_backfilled', 0)}",
                "green" if _match_pct >= 95 else "amber",
            )
            yield event(
                "  →", "TIMING", f"DC parse: {_time.monotonic() - _t:.1f}s", "muted"
            )
        except Exception as e:
            yield event("  →", "ERROR", f"DC pipeline failed: {str(e)[:100]}", "red")
        await asyncio.sleep(0.1)

        # ── Step 4: Verify XSUP + COE in DB ─────────────────────────────
        _t = _time.monotonic()
        yield event(
            "4/5 DB Verification",
            "CHECK",
            "Confirming XSUP Tracker · COE Tracker · Data Quality",
        )
        try:
            with get_db() as _vc:
                # XSUP
                _xt = _vc.execute("SELECT COUNT(*) FROM xsup_data").fetchone()[0]
                _xp1 = _vc.execute(
                    "SELECT COUNT(*) FROM xsup_data WHERE xsup_priority='P1'"
                ).fetchone()[0]
                _xp2 = _vc.execute(
                    "SELECT COUNT(*) FROM xsup_data WHERE xsup_priority='P2'"
                ).fetchone()[0]
                _xm = _vc.execute(
                    "SELECT COUNT(*) FROM xsup_data WHERE account_id IS NOT NULL"
                ).fetchone()[0]
                _xsync = _vc.execute(
                    "SELECT synced_at FROM xsup_data ORDER BY id DESC LIMIT 1"
                ).fetchone()
                _xts = (_xsync["synced_at"] or "")[:16] if _xsync else "never"
                _xmatch = round(100 * _xm / _xt) if _xt else 0
                yield event(
                    "  → XSUP",
                    "OK" if _xt > 0 else "WARN",
                    f"{_xt} open XSUPs · {_xp1} P1 · {_xp2} P2 · "
                    f"{_xm}/{_xt} matched to accounts ({_xmatch}%) · synced {_xts}",
                    "green" if _xt > 100 else "amber",
                )
                # COE Issues
                _ci = _vc.execute("SELECT COUNT(*) FROM coe_issues").fetchone()[0]
                _cim = _vc.execute(
                    "SELECT COUNT(*) FROM coe_issues WHERE account_id IS NOT NULL"
                ).fetchone()[0]
                _cb = _vc.execute("SELECT COUNT(*) FROM coe_bugs").fetchone()[0]
                _cbm = _vc.execute(
                    "SELECT COUNT(*) FROM coe_bugs WHERE account_id IS NOT NULL"
                ).fetchone()[0]
                _cisync = _vc.execute(
                    "SELECT synced_at FROM coe_issues ORDER BY rowid DESC LIMIT 1"
                ).fetchone()
                _cts = (_cisync["synced_at"] or "")[:16] if _cisync else "never"
                _cimatch = round(100 * _cim / _ci) if _ci else 0
                _cbmatch = round(100 * _cbm / _cb) if _cb else 0
                yield event(
                    "  → COE Issues",
                    "OK" if _ci > 0 else "WARN",
                    f"{_ci} issues · {_cim}/{_ci} matched ({_cimatch}%) · synced {_cts}",
                    "green" if _ci > 100 else "amber",
                )
                yield event(
                    "  → COE Bugs",
                    "OK" if _cb > 0 else "WARN",
                    f"{_cb} bugs · {_cbm}/{_cb} matched ({_cbmatch}%)",
                    "green" if _cb > 100 else "amber",
                )
                # Accounts + CSE coverage
                _n = _vc.execute("SELECT COUNT(*) FROM accounts").fetchone()[0]
                _no_cse = _vc.execute(
                    "SELECT COUNT(*) FROM accounts WHERE active_cse IS NULL OR active_cse=''"
                ).fetchone()[0]
                _m9 = _vc.execute(
                    "SELECT COALESCE(SUM(m9_complete),0) FROM blocked_data WHERE cohort='Scale cohort'"
                ).fetchone()[0]
                _m8 = _vc.execute(
                    "SELECT COUNT(*) FROM blocked_data WHERE m8_started=1 AND m9_complete=0 AND cohort='Scale cohort'"
                ).fetchone()[0]
                _cse_pct = round(100 * (_n - _no_cse) / _n) if _n else 0
                yield event(
                    "  → Accounts",
                    "OK",
                    f"{_n} accounts · CSE coverage {_n - _no_cse}/{_n} ({_cse_pct}%) · "
                    f"M8 in-flight: {_m8} · M9 complete: {_m9}",
                    "green" if _cse_pct >= 70 else "amber",
                )
        except Exception as e:
            yield event("  →", "ERROR", f"DB verify failed: {str(e)[:100]}", "red")
        yield event(
            "  →", "TIMING", f"DB verify: {_time.monotonic() - _t:.1f}s", "muted"
        )
        await asyncio.sleep(0.1)

        # ── Step 5: Summary ───────────────────────────────────────────────
        _total = round(_time.monotonic() - _pipeline_start, 1)
        yield event(
            "5/5 Complete",
            "DONE",
            f"Pipeline finished in {_total}s — dashboard refreshing",
            "teal",
        )
        yield f"data: {_j.dumps({'done': True})}\n\n"

    return StreamingResponse(stream(), media_type="text/event-stream")


def _load_m9_schedule(cohort: str = "") -> list:
    _ensure_db()
    """M9 completion schedule — M3 complete, M9 planned, not yet complete, sorted by date."""
    try:
        from datetime import datetime, date

        today = date.today()
        with get_db() as conn:
            rows = conn.execute(f"""
                SELECT a.customer_name, a.active_cse, a.status, a.live_fire, a.live_fire_dc,
                       b.m9_planned, b.m8_started, b.m9_complete
                FROM accounts a JOIN blocked_data b ON a.account_id=b.account_id
                WHERE b.m3_complete=1 AND b.m9_planned!='' AND b.m9_complete=0
                  {_cohort_sql()}
                ORDER BY b.m9_planned, a.customer_name
            """, (cohort, cohort)).fetchall()
        result = []
        for r in rows:
            try:
                d = None
                for fmt in ("%m/%d/%Y", "%m/%d/%y"):
                    try:
                        d = datetime.strptime(r["m9_planned"].strip(), fmt).date()
                        break
                    except:
                        pass
                if d and d >= today:
                    result.append(
                        {
                            "name": r["customer_name"],
                            "cse": r["active_cse"] or "—",
                            "status": r["status"] or "—",
                            "m9_planned": r["m9_planned"],
                            "m9_date_iso": d.isoformat(),
                            "m9_date": d.strftime("%d %b %Y"),
                            "m8_started": bool(r["m8_started"]),
                            "live_fire": bool(r["live_fire"]),
                            "live_fire_dc": r["live_fire_dc"] or "",
                        }
                    )
            except:
                pass
        return result
    except:
        return []


@app.get("/api/m9-schedule")
def api_m9(cohort: str = ""):
    return _load_m9_schedule(cohort=cohort)


def _load_in_progress() -> list:
    _ensure_db()
    """Upgrades actively in progress — M8 started, M9 not complete."""
    try:
        with get_db() as conn:
            rows = conn.execute("""
                SELECT a.customer_name, a.active_cse, a.status, a.live_fire, a.live_fire_dc,
                       b.m8_planned, b.m9_planned, b.upgrade_notes, b.health_notes,
                       b.signal, b.subtype, b.dc_progress, b.churn_risk, b.cc_rep
                FROM accounts a JOIN blocked_data b ON a.account_id=b.account_id
                WHERE b.m8_started=1 AND b.m9_complete=0
                ORDER BY b.m9_planned, a.customer_name
            """).fetchall()
        return [dict(r) for r in rows]
    except:
        return []


@app.get("/api/in-progress")
def api_in_progress():
    return _load_in_progress()


def _load_open_actions() -> list:
    _ensure_db()
    """Open Actions — accounts needing follow-up grouped by category."""
    try:
        import json as _j

        state = _j.loads(STATE_FILE.read_text())
        groups = {
            "Sales Hold": {"statuses": ["Sales Hold"], "color": "#EA580C"},
            "Churning / Churned": {
                "statuses": ["Churning/Churned"],
                "color": "#DC2626",
            },
            "Ready To Engage": {"statuses": ["Ready To Engage"], "color": "#10B981"},
            "Account Contacted": {
                "statuses": ["Account team contacted"],
                "color": "#F59E0B",
            },
            "Blocked": {"statuses": ["Blocked: Tech limitation"], "color": "#A1887F"},
            "On Hold": {"statuses": ["On Hold"], "color": "#3B82F6"},
            "Escalation": {"statuses": ["Backoff", "Cancelled"], "color": "#EF4444"},
        }
        result = []
        for gname, cfg in groups.items():
            accs = []
            for acc in state.get("accounts", {}).values():
                if acc.get("status") not in cfg["statuses"]:
                    continue
                if not acc.get("customer_name", "").strip():
                    continue
                bd = acc.get("blocked_data") or {}
                ai = acc.get("ai_enrichment") or {}
                accs.append(
                    {
                        "name": acc.get("customer_name", "—"),
                        "cse": acc.get("active_cse") or "—",
                        "region": acc.get("sales_region") or "—",
                        "status": acc.get("status", "—"),
                        "signal": bd.get("signal", ""),
                        "blocker": ai.get("blocker", ""),
                        "accountable": ai.get("accountable", ""),
                        "live_fire": acc.get("live_fire", False),
                        "live_fire_dc": acc.get("live_fire_dc", ""),
                    }
                )
            if accs:
                result.append(
                    {
                        "group": gname,
                        "color": cfg["color"],
                        "accounts": sorted(accs, key=lambda x: x["name"]),
                    }
                )
        return result
    except Exception as e:
        return []


def _load_milestones(theatre: str = "", cohort: str = "") -> list:
    _ensure_db()
    """Milestone tracker — full M0-M9 with SLA breach flags. Optional theatre filter."""
    from datetime import datetime

    MARCH9 = datetime(2026, 3, 9)

    def _pd(s):
        if not s:
            return None
        for f in ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%Y %H:%M:%S"):
            try:
                return datetime.strptime(
                    s.strip().split(" ")[0] if " " in s else s.strip(), f
                )
            except:
                pass
        return None

    try:
        with get_db() as conn:
            rows = conn.execute(
                f"""
                SELECT a.account_id, a.customer_name, a.active_cse, a.status, a.live_fire, a.live_fire_dc,
                       a.sales_region,
                       b.team, b.is_cs_team, b.signal, b.subtype, b.milestone_category,
                       b.m0_complete, b.m1_complete, b.m1_planned,
                       b.m2_complete, b.m2_planned,
                       b.m3_complete, b.m3_planned,
                       b.m4_complete, b.m5_complete,
                       b.m7_complete,
                       b.m8_started, b.m8_planned, b.m8_actual,
                       b.m9_complete, b.m9_planned, b.m9_actual,
                       b.upgrade_notes, b.health_notes, b.exec_delay, b.status_detail,
                       b.dc_progress, b.owner_e2e,
                       COALESCE(a.account_theatre, b.account_theatre, 'EMEA') as account_theatre
                FROM accounts a JOIN blocked_data b ON a.account_id=b.account_id
                WHERE a.customer_name != ''
                  {_cohort_sql()}
                  AND (? = '' OR UPPER(COALESCE(a.account_theatre,'EMEA')) = UPPER(?))
                ORDER BY b.signal, a.customer_name
            """,
                (cohort, cohort, theatre, theatre),
            ).fetchall()
            result = []
            for r in rows:
                d = dict(r)
                m3d = _pd(d.get("m3_planned"))
                m8d = _pd(d.get("m8_planned"))
                m9d = _pd(d.get("m9_planned"))
                # M3→M8: only penalise if M3 >= Mar 9 (prospective)
                # M8→M9: only penalise if M8 >= Mar 9 (regardless of M3 date)
                sla_m3_m8 = None
                sla_m8_m9 = None
                if m3d and m3d >= MARCH9 and m8d:
                    sla_m3_m8 = (m8d - m3d).days
                if m8d and m8d >= MARCH9 and m9d:
                    sla_m8_m9 = (m9d - m8d).days
                d["sla_m3_m8_days"] = sla_m3_m8
                d["sla_m8_m9_days"] = sla_m8_m9
                d["sla_m3_m8_breach"] = sla_m3_m8 is not None and sla_m3_m8 > 14
                d["sla_m8_m9_breach"] = sla_m8_m9 is not None and sla_m8_m9 > 28
                result.append(d)
            return result
    except:
        return []


def _load_ps() -> dict:
    _ensure_db()
    """PS engagement — matched + unmatched."""
    try:
        with get_db() as conn:
            matched = conn.execute("""
                SELECT a.customer_name, a.active_cse, a.live_fire, a.live_fire_dc,
                       p.psc, p.psc_shadow, p.pm, p.ps_status, p.clarizen_id, p.timeline, p.notes
                FROM accounts a JOIN ps_data p ON a.account_id=p.account_id
                ORDER BY p.ps_status, a.customer_name
            """).fetchall()
        import csv as _csv

        ps_file = DATA_DIR / "ps_tracker.csv"
        all_ps = (
            list(_csv.DictReader(open(ps_file, encoding="utf-8-sig")))
            if ps_file.exists()
            else []
        )
        with get_db() as conn2:
            matched_ps_names = {
                r[0] for r in conn2.execute("SELECT ps_name FROM ps_data").fetchall()
            }
        unmatched = [
            {
                "name": r["PS Eligible Account Name"],
                "country": r.get("Theater", ""),
                "psc": r.get("Area Owner", ""),
                "timeline": r.get("Number of Sessions (4 hours)", ""),
            }
            for r in all_ps
            if r.get("PS Eligible Account Name", "").strip()
            and r.get("PS Eligible Account Name", "").strip() not in matched_ps_names
        ]
        return {
            "matched": [dict(r) for r in matched],
            "unmatched": sorted(unmatched, key=lambda x: x["name"]),
        }
    except:
        return {"matched": [], "unmatched": []}


def _load_completed(theatre: str = "", cohort: str = "") -> list:
    _ensure_db()
    """Completed accounts — DC M9 complete is source of truth."""
    try:
        with get_db() as conn:
            rows = conn.execute(
                f"""
                SELECT a.customer_name, a.active_cse, a.sales_region,
                       a.live_fire, a.live_fire_dc,
                       b.m9_actual, b.m9_planned, b.dc_progress
                FROM accounts a
                JOIN blocked_data b ON a.account_id=b.account_id
                WHERE b.m9_complete=1
                  {_cohort_sql()}
                  AND (? = '' OR UPPER(COALESCE(a.account_theatre,'EMEA'))=UPPER(?))
                ORDER BY b.m9_actual DESC, b.m9_planned DESC
            """,
                (cohort, cohort, theatre, theatre),
            ).fetchall()
        return [dict(r) for r in rows]
    except Exception as e:
        return []


def _load_dq() -> list:
    _ensure_db()
    """Data quality issues — DC CSE Tracker is source of truth."""
    try:
        with get_db() as conn:
            # Source of truth: DB (populated from DC tracker)
            rows = conn.execute("""
                SELECT a.account_id, a.customer_name, a.active_cse, a.status,
                       a.email_sent, b.dc_progress, b.m1_complete
                FROM accounts a
                LEFT JOIN blocked_data b ON a.account_id=b.account_id
                WHERE a.customer_name!=''
            """).fetchall()
        accs = [dict(r) for r in rows]
        OUTREACH = {"Ready To Engage", "Account team contacted", "Upgrade Email Sent"}
        TERMINAL = {"Churning/Churned", "Cancelled", "Backoff", "Completed"}
        SKIP_CSE = TERMINAL | {"PS"}
        issues = []
        # No status: only flag if DC also has no meaningful status
        no_st = [a for a in accs if not (a.get("status") or "").strip()]
        no_cse = [
            a
            for a in accs
            if not (a.get("active_cse") or "").strip()
            and (a.get("status") or "").strip() not in SKIP_CSE
        ]
        no_email = [
            a
            for a in accs
            if (a.get("status") or "") in OUTREACH
            and not (a.get("email_sent") or "").strip()
        ]
        if no_st:
            issues.append(
                {
                    "type": "No Status",
                    "count": len(no_st),
                    "accounts": [a["customer_name"] for a in no_st],
                }
            )
        if no_cse:
            issues.append(
                {
                    "type": "No Owner/CSE",
                    "count": len(no_cse),
                    "accounts": [a["customer_name"] for a in no_cse],
                }
            )
        if no_email:
            issues.append(
                {
                    "type": "No Email on Record",
                    "count": len(no_email),
                    "accounts": [a["customer_name"] for a in no_email],
                }
            )
        return issues
    except:
        return []


def _load_audit_log(theatre: str = "") -> list:
    _ensure_db()
    """Audit log — changes detected across all source files between pipeline runs."""
    try:
        with get_db() as conn:
            rows = conn.execute(
                """
                SELECT
                  CASE WHEN sh.account_id='unmatched_dc' THEN sh.new_status
                       ELSE a.customer_name END as customer_name,
                  CASE WHEN sh.account_id='unmatched_dc' THEN sh.old_status
                       ELSE COALESCE(a.active_cse,'—') END as active_cse,
                  CASE WHEN sh.account_id='unmatched_dc' THEN 'N' ELSE sh.old_status END as old_status,
                  CASE WHEN sh.account_id='unmatched_dc' THEN 'Y' ELSE sh.new_status END as new_status,
                  sh.changed_at, sh.source,
                  COALESCE(sh.file_source, 'DC CSE Tracker') as file_source,
                  COALESCE(sh.field_name, 'status') as field_name
                FROM status_history sh
                LEFT JOIN accounts a ON a.account_id = sh.account_id
                WHERE sh.source IN ('pipeline','backfill')
                  AND sh.file_source = 'DC CSE Tracker'
                  AND (? = ''
                       OR UPPER(COALESCE(a.account_theatre,'EMEA'))=UPPER(?)
                       OR (sh.account_id='unmatched_dc' AND (
                           ? = '' OR UPPER(COALESCE(
                             (SELECT a2.account_theatre FROM accounts a2
                              WHERE a2.customer_name=sh.new_status LIMIT 1),
                             'EMEA'))=UPPER(?))))
                ORDER BY sh.changed_at DESC
                LIMIT 2000
            """,
                (theatre, theatre, theatre, theatre),
            ).fetchall()
        return [dict(r) for r in rows]
    except:
        return []


@app.get("/api/audit-log")
def api_audit(theatre: str = ""):
    return _load_audit_log(theatre=theatre)


@app.get("/api/update-blocked-tab")
def api_update_tab(url: str):
    """
    Update the blocked accounts tab gid from a Google Sheets URL.
    Usage: /api/update-blocked-tab?url=https://docs.google.com/...gid=538753662
    """
    import re, json as _j

    m = re.search(r"gid=(\d+)", url)
    if not m:
        return {"status": "error", "message": "No gid found in URL"}
    gid = m.group(1)
    config_path = DATA_DIR / "drive_config.json"
    config = _j.loads(config_path.read_text())
    updated = False
    for f in config["files"]:
        if "Blocked" in f["name"]:
            old_gid = f.get("current_gid", "—")
            f["current_gid"] = gid
            updated = True
            break
    if updated:
        config_path.write_text(_j.dumps(config, indent=2))
        return {
            "status": "ok",
            "message": f"Blocked accounts tab updated → gid={gid}",
            "old_gid": old_gid,
            "new_gid": gid,
        }
    return {"status": "error", "message": "Blocked accounts file not found in config"}


@app.get("/api/sla-breaches")
def api_sla_breaches(theatre: str = ""):
    """SLA breach report — M3→M8 >14d or M8→M9 >28d, prospective from Mar 9."""
    milestones = _load_milestones(theatre=theatre)
    breaches = [
        {
            k: v
            for k, v in r.items()
            if k
            in (
                "account_id",
                "customer_name",
                "active_cse",
                "status",
                "live_fire",
                "dc_progress",
                "m3_planned",
                "m8_planned",
                "m9_planned",
                "m3_complete",
                "m8_started",
                "m9_complete",
                "sla_m3_m8_days",
                "sla_m8_m9_days",
                "sla_m3_m8_breach",
                "sla_m8_m9_breach",
                "status_detail",
                "owner_e2e",
            )
        }
        for r in milestones
        if r.get("sla_m3_m8_breach") or r.get("sla_m8_m9_breach")
    ]
    return sorted(
        breaches,
        key=lambda x: -(x.get("sla_m3_m8_days") or 0) - (x.get("sla_m8_m9_days") or 0),
    )


def _get_blocked_export_url() -> str:
    """
    Get CSV export URL for blocked accounts — always uses stored gid.
    Tab format: EMEA_MONTH DAY (e.g. EMEA_MARCH 22, EMEA_APRIL 5).
    gid must be updated in drive_config.json when tab changes.
    Use /api/update-blocked-tab?url=<google_sheets_url> to update.
    """
    import json as _j

    config = _j.loads((DATA_DIR / "drive_config.json").read_text())
    for f in config["files"]:
        if "Blocked" in f["name"]:
            file_id = f.get("file_id", "")
            gid = f.get("current_gid", "0")
            return f"https://docs.google.com/spreadsheets/d/{file_id}/export?format=csv&gid={gid}"
    return ""


@app.get("/api/open-actions")
def api_open_actions():
    return _load_open_actions()


@app.get("/api/milestones")
def api_milestones(theatre: str = "", cohort: str = ""):
    return _load_milestones(theatre=theatre, cohort=cohort)


@app.get("/api/health-summary")
def api_health_summary(cohort: str = ""):
    """Theatre health status — green/amber/red per theatre based on blocked count."""
    _ensure_db()
    theatres = ["EMEA", "JAPAC", "AMER", "LATAM"]
    result = {}
    try:
        with get_db() as conn:
            for theatre in theatres:
                rows = conn.execute(
                    f"""
                    SELECT b.signal, b.m9_complete
                    FROM blocked_data b
                    JOIN accounts a ON a.account_id=b.account_id
                    WHERE UPPER(COALESCE(b.account_theatre, a.account_theatre,'EMEA'))=?
                      AND a.customer_name!=''
                      {_cohort_sql()}
                """,
                    (theatre, cohort, cohort),
                ).fetchall()
                m9 = sum(1 for r in rows if r[1])
                blocked = sum(1 for r in rows if r[0] == "blocked" and not r[1])
                at_risk = sum(1 for r in rows if r[0] == "at_risk" and not r[1])
                if blocked > 5:
                    status = "red"
                elif blocked > 2:
                    status = "amber"
                else:
                    status = "green"
                result[theatre] = {
                    "status": status,
                    "m9": m9,
                    "blocked": blocked,
                    "at_risk": at_risk,
                }
    except Exception as e:
        logger.error("health-summary failed: %s", e)
        for t in theatres:
            result[t] = {
                "status": "amber",
                "m9": 0,
                "blocked": 0,
                "at_risk": 0,
                "error": True,
            }
    return result


@app.get("/api/cse-workload")
def api_cse_workload(theatre: str = "", cohort: str = ""):
    """Per-CSE account load, blocked/at-risk counts, M9 this month."""
    from datetime import date as _date, datetime as _dt

    _ensure_db()
    try:
        today = _date.today()
        cur_year, cur_month = today.year, today.month

        def _is_this_month(s):
            if not s:
                return False
            s = str(s).strip()
            for fmt in (
                "%m/%d/%Y %H:%M:%S",
                "%m/%d/%Y",
                "%Y-%m-%d %H:%M:%S",
                "%Y-%m-%d",
            ):
                try:
                    d = _dt.strptime(
                        s.split(" ")[0] if " " in s and not fmt.endswith("%S") else s,
                        fmt.split(" ")[0],
                    ).date()
                    return d.year == cur_year and d.month == cur_month
                except Exception:
                    pass
            return False

        with get_db() as conn:
            rows = conn.execute(
                f"""
                SELECT a.active_cse as cse,
                       COUNT(*) as account_count,
                       SUM(CASE WHEN b.signal='blocked' AND b.m9_complete=0 THEN 1 ELSE 0 END) as blocked_count,
                       SUM(CASE WHEN b.signal='at_risk' AND b.m9_complete=0 THEN 1 ELSE 0 END) as at_risk_count,
                       SUM(CASE WHEN b.m8_started=1 AND b.m9_complete=0 THEN 1 ELSE 0 END) as m8_count,
                       b.m9_complete, b.m9_actual
                FROM accounts a
                JOIN blocked_data b ON a.account_id=b.account_id
                WHERE a.active_cse!='' AND a.customer_name!=''
                  {_cohort_sql()}
                  AND (? = '' OR UPPER(COALESCE(b.account_theatre, a.account_theatre,'EMEA'))=UPPER(?))
                GROUP BY a.active_cse, b.m9_complete, b.m9_actual
                ORDER BY a.active_cse
            """,
                (cohort, cohort, theatre, theatre),
            ).fetchall()

        # Aggregate in Python so DC date parsing works correctly
        from collections import defaultdict

        agg = defaultdict(
            lambda: {
                "account_count": 0,
                "blocked_count": 0,
                "at_risk_count": 0,
                "m9_this_month": 0,
                "m8_count": 0,
            }
        )
        for r in rows:
            cse = r["cse"]
            agg[cse]["account_count"] += r["account_count"]
            agg[cse]["blocked_count"] += r["blocked_count"]
            agg[cse]["at_risk_count"] += r["at_risk_count"]
            agg[cse]["m8_count"] += r["m8_count"]
            if r["m9_complete"] and _is_this_month(r["m9_actual"]):
                agg[cse]["m9_this_month"] += r["account_count"]

        result = [{"cse": cse, **v} for cse, v in agg.items()]
        result.sort(key=lambda x: (-x["m8_count"], -x["account_count"]))
        return result
    except Exception as e:
        logger.error("cse-workload failed: %s", e)
        return []


@app.get("/api/weekly-movements")
def api_weekly_movements(theatre: str = "", date: str = "", cohort: str = ""):
    """What changed this week — new M9, M8 started, newly blocked, resolved."""
    from datetime import date as _date, timedelta

    _ensure_db()
    try:
        if date:
            ref = _date.fromisoformat(date)
        else:
            ref = _date.today()
        monday = ref - timedelta(days=ref.weekday())
        sunday = monday + timedelta(days=6)
        mon_s = monday.isoformat()
        sun_s = sunday.isoformat()

        t_filter = (
            "AND UPPER(COALESCE(b.account_theatre, a.account_theatre,'EMEA'))=UPPER(?)"
            if theatre
            else ""
        )
        t_params = (cohort, cohort, theatre) if theatre else (cohort, cohort)

        def _q(sql, params):
            with get_db() as conn:
                return [dict(r) for r in conn.execute(sql, params).fetchall()]

        def _pd(s):
            """Parse DC date strings in any of the 3 known formats."""
            if not s:
                return None
            s = str(s).strip()
            for fmt in (
                "%m/%d/%Y %H:%M:%S",
                "%m/%d/%Y",
                "%Y-%m-%d %H:%M:%S",
                "%Y-%m-%d",
            ):
                try:
                    from datetime import datetime as _dt

                    return _dt.strptime(
                        s.split(" ")[0] if " " in s and fmt.endswith("%S") else s,
                        fmt.split(" ")[0],
                    ).date()
                except:
                    pass
            return None

        base = f"""
            FROM accounts a JOIN blocked_data b ON a.account_id=b.account_id
            WHERE a.customer_name!='' {_cohort_sql()} {t_filter}
        """

        def _sh_milestone(field_name):
            """Accounts that hit a milestone this week via status_history."""
            return _q(
                f"""
                SELECT a.account_id, a.customer_name, a.active_cse, a.sales_region,
                       MAX(sh.changed_at) as milestone_date,
                       COALESCE(b.account_theatre, a.account_theatre,'EMEA') as account_theatre
                FROM accounts a
                JOIN blocked_data b ON a.account_id=b.account_id
                JOIN status_history sh ON sh.account_id=a.account_id
                WHERE a.customer_name!='' {_cohort_sql()} {t_filter}
                  AND sh.field_name=? AND sh.new_status='Y'
                  AND sh.changed_at>=? AND sh.changed_at<=?
                GROUP BY a.account_id
                ORDER BY milestone_date DESC
                """,
                t_params + (field_name, mon_s, sun_s + "T23:59:59"),
            )

        # Fetch all M9/M8 completed and filter by date in Python (DC dates are M/D/YYYY not ISO)
        all_m9 = _q(
            f"""
            SELECT a.account_id, a.customer_name, a.active_cse, a.sales_region,
                   b.m9_actual, b.dc_progress,
                   COALESCE(b.account_theatre, a.account_theatre,'EMEA') as account_theatre
            {base} AND b.m9_complete=1 AND b.m9_actual!=''
            ORDER BY b.m9_actual DESC
        """,
            t_params,
        )
        new_m9 = [
            r
            for r in all_m9
            if monday <= (_pd(r.get("m9_actual")) or _date.min) <= sunday
        ]

        all_m8 = _q(
            f"""
            SELECT a.account_id, a.customer_name, a.active_cse, a.sales_region,
                   b.m8_actual,
                   COALESCE(b.account_theatre, a.account_theatre,'EMEA') as account_theatre
            {base} AND b.m8_started=1 AND b.m8_actual!=''
            ORDER BY b.m8_actual DESC
        """,
            t_params,
        )
        m8_started = [
            r
            for r in all_m8
            if monday <= (_pd(r.get("m8_actual")) or _date.min) <= sunday
        ]

        newly_blocked = _q(
            f"""
            SELECT a.account_id, a.customer_name, a.active_cse, a.sales_region,
                   b.signal, b.subtype,
                   COALESCE(b.account_theatre, a.account_theatre,'EMEA') as account_theatre
            {base} AND b.signal IN ('blocked','at_risk') AND b.m9_complete=0
              AND EXISTS (
                SELECT 1 FROM status_history sh
                WHERE sh.account_id=a.account_id
                  AND sh.field_name='signal'
                  AND sh.new_status IN ('blocked','at_risk')
                  AND sh.changed_at>=? AND sh.changed_at<=?
              )
            ORDER BY a.customer_name
        """,
            t_params + (mon_s, sun_s + "T23:59:59"),
        )

        resolved = _q(
            f"""
            SELECT a.account_id, a.customer_name, a.active_cse, a.sales_region,
                   COALESCE(b.account_theatre, a.account_theatre,'EMEA') as account_theatre
            {base} AND b.m9_complete=0 AND b.signal='green'
              AND EXISTS (
                SELECT 1 FROM status_history sh
                WHERE sh.account_id=a.account_id
                  AND sh.field_name='signal'
                  AND sh.old_status IN ('blocked','at_risk')
                  AND sh.new_status='green'
                  AND sh.changed_at>=? AND sh.changed_at<=?
              )
            ORDER BY a.customer_name
        """,
            t_params + (mon_s, sun_s + "T23:59:59"),
        )

        return {
            "week_of": mon_s,
            "m1_outreach": _sh_milestone("M1 Outreach"),
            "m2_entitlements": _sh_milestone("M2 Entitlements"),
            "m3_buyin": _sh_milestone("M3 Buy-in"),
            "m4_discovery": _sh_milestone("M4 Discovery"),
            "m5_tech": _sh_milestone("M5 Tech Validation"),
            "m8_started": m8_started,
            "new_m9": new_m9,
            "newly_blocked": newly_blocked,
            "resolved": resolved,
        }
    except Exception as e:
        logger.error("weekly-movements failed: %s", e)
        return {
            "week_of": "",
            "new_m9": [],
            "m8_started": [],
            "newly_blocked": [],
            "resolved": [],
        }


@app.get("/api/compare")
def api_compare(cohort: str = ""):
    """4-theatre side-by-side comparison for QBR."""
    from datetime import date as _date, timedelta

    _ensure_db()
    theatres = ["EMEA", "JAPAC", "AMER", "LATAM"]
    result = []
    today = _date.today()
    monday = (today - timedelta(days=today.weekday())).isoformat()
    try:
        with get_db() as conn:
            for theatre in theatres:
                rows = conn.execute(
                    f"""
                    SELECT b.signal, b.m9_complete, b.m9_actual
                    FROM blocked_data b
                    JOIN accounts a ON a.account_id=b.account_id
                    WHERE UPPER(COALESCE(b.account_theatre, a.account_theatre,'EMEA'))=?
                      AND a.customer_name!=''
                      {_cohort_sql()}
                """,
                    (theatre, cohort, cohort),
                ).fetchall()
                m9_total = sum(1 for r in rows if r[1])
                m9_this_week = sum(1 for r in rows if r[1] and r[2] and r[2] >= monday)
                blocked = sum(1 for r in rows if r[0] == "blocked" and not r[1])
                at_risk = sum(1 for r in rows if r[0] == "at_risk" and not r[1])
                sla_rows = _load_milestones(theatre=theatre, cohort=cohort)
                sla_overdue = sum(
                    1
                    for r in sla_rows
                    if r.get("sla_m3_m8_breach") or r.get("sla_m8_m9_breach")
                )
                result.append(
                    {
                        "theatre": theatre,
                        "m9_total": m9_total,
                        "m9_this_week": m9_this_week,
                        "blocked": blocked,
                        "at_risk": at_risk,
                        "sla_overdue": sla_overdue,
                    }
                )
    except Exception as e:
        logger.error("compare failed: %s", e)
        for t in theatres:
            result.append(
                {
                    "theatre": t,
                    "m9_total": 0,
                    "m9_this_week": 0,
                    "blocked": 0,
                    "at_risk": 0,
                    "sla_overdue": 0,
                }
            )
    return {"theatres": result}


@app.get("/api/theatres")
def api_theatres():
    """List distinct theatres with account counts."""
    _ensure_db()
    try:
        with get_db() as conn:
            rows = conn.execute("""
                SELECT COALESCE(account_theatre,'EMEA') theatre, COUNT(*) n
                FROM accounts WHERE customer_name!=''
                GROUP BY account_theatre ORDER BY n DESC
            """).fetchall()
        return [dict(r) for r in rows]
    except:
        return []


@app.get("/api/ps")
def api_ps():
    return _load_ps()


@app.get("/api/completed")
def api_completed(theatre: str = "", cohort: str = ""):
    return _load_completed(theatre=theatre, cohort=cohort)


@app.get("/api/blockers")
def api_blockers(theatre: str = "", region: str = "", cse: str = "", cohort: str = ""):
    """Blocked accounts grouped by blocker type for call prep."""
    _ensure_db()
    try:
        with get_db() as conn:
            rows = conn.execute(
                f"""
                SELECT a.customer_name, a.active_cse, a.sales_region,
                       COALESCE(a.account_theatre,'EMEA') as account_theatre,
                       b.signal, b.subtype, b.status_detail, b.upgrade_notes,
                       b.health_notes, b.dc_progress, b.churn_risk,
                       b.cc_rep, b.cc_dsm, b.cohort, b.area, b.district,
                       b.m8_started, b.m9_complete, b.m3_complete,
                       b.last_edited_by, b.last_edited_date,
                       b.current_project_status,
                       b.m1_details, b.m3_details, b.m5_details,
                       a.account_id,
                       b.m0_complete, b.m1_complete, b.m1_planned,
                       b.m2_complete, b.m2_planned, b.m3_planned,
                       b.m4_complete, b.m4_planned,
                       b.m5_complete, b.m5_planned,
                       b.m6_complete, b.m7_complete, b.m7_planned,
                       b.m8_planned, b.m9_planned, b.m9_actual
                FROM accounts a JOIN blocked_data b ON a.account_id=b.account_id
                WHERE a.customer_name != ''
                  AND b.signal IN ('blocked','at_risk')
                  AND b.m9_complete = 0
                  {_cohort_sql()}
                  AND (? = '' OR UPPER(COALESCE(a.account_theatre,'EMEA'))=UPPER(?))
                  AND (? = '' OR LOWER(a.sales_region) LIKE LOWER(?))
                  AND (? = '' OR a.active_cse = ?)
                ORDER BY b.subtype, a.sales_region, a.customer_name
            """,
                (cohort, cohort, theatre, theatre, region, f"%{region}%", cse, cse),
            ).fetchall()
        known = {
            "churn",
            "no_contact",
            "customer_delay",
            "core_rep_blocking",
            "tech_blocker",
            "active_deal",
            "legal_blocker",
            "self_hosted",
        }
        from agent.dc_parser import _subtype_from_detail

        result: dict = {}
        for r in rows:
            d = dict(r)
            st = d.get("subtype") or ""
            # Re-classify on-the-fly if not in known set (handles stale DB entries)
            if st not in known:
                st = _subtype_from_detail(d.get("status_detail") or "") or "other"
            bucket = st if st in known else "other"
            result.setdefault(bucket, []).append(d)

        # Enrich every account with COE issues + bugs (one bulk query each)
        all_ids = list(
            {
                d["account_id"]
                for accounts in result.values()
                for d in accounts
                if d.get("account_id")
            }
        )
        if all_ids:
            placeholders = ",".join("?" * len(all_ids))
            with get_db() as conn2:
                coe_issues_rows = conn2.execute(
                    f"""SELECT account_id, issue_id, upgrade_blocker, technical_issue,
                               priority, module, issue_category, status, timeline_answer
                        FROM coe_issues WHERE account_id IN ({placeholders})
                        ORDER BY account_id, priority, issue_id""",
                    all_ids,
                ).fetchall()
                coe_bugs_rows = conn2.execute(
                    f"""SELECT account_id, xsup_number, xsup_priority, xsup_status,
                               spo_dc_classification, eng_escalation_status, component, summary
                        FROM coe_bugs WHERE account_id IN ({placeholders})
                        ORDER BY account_id, xsup_priority, xsup_number""",
                    all_ids,
                ).fetchall()

            issues_by_acct: dict = {}
            for r in coe_issues_rows:
                issues_by_acct.setdefault(r["account_id"], []).append(dict(r))
            bugs_by_acct: dict = {}
            for r in coe_bugs_rows:
                bugs_by_acct.setdefault(r["account_id"], []).append(dict(r))

            for accounts in result.values():
                for d in accounts:
                    aid = d.get("account_id") or ""
                    d["coe_issues"] = issues_by_acct.get(aid, [])
                    d["coe_bugs"] = bugs_by_acct.get(aid, [])

        return result
    except Exception as e:
        return {"error": str(e)}


@app.get("/api/forecast")
def api_forecast(theatre: str = "", cohort: str = ""):
    """Next 7 days M8/M9 targets + 4-week velocity."""
    _ensure_db()
    from datetime import datetime, timedelta

    today = datetime.now(timezone.utc).date()
    next_week_end = today + timedelta(days=7)
    three_month_end = today + timedelta(days=90)

    try:
        with get_db() as conn:
            targets = conn.execute(
                f"""
                SELECT a.account_id, a.customer_name, a.active_cse, a.sales_region,
                       COALESCE(a.account_theatre,'EMEA') as account_theatre,
                       b.m9_planned, b.m8_planned, b.m8_started, b.m9_complete,
                       b.dc_progress, b.churn_risk, b.m8_actual,
                       b.subtype, b.signal
                FROM accounts a JOIN blocked_data b ON a.account_id=b.account_id
                WHERE a.customer_name != '' AND b.m9_complete=0
                  {_cohort_sql()}
                  AND (? = '' OR UPPER(COALESCE(a.account_theatre,'EMEA'))=UPPER(?))
                ORDER BY b.m9_planned
            """,
                (cohort, cohort, theatre, theatre),
            ).fetchall()

            def _pdw(s):
                if not s or s == "None":
                    return None
                s = s.strip()
                for fmt in (
                    "%m/%d/%Y %H:%M:%S",
                    "%m/%d/%Y",
                    "%Y-%m-%d %H:%M:%S",
                    "%Y-%m-%d",
                ):
                    try:
                        from datetime import datetime as _dt2

                        return _dt2.strptime(
                            s.split(" ")[0] if " " in s else s, fmt.split(" ")[0]
                        ).date()
                    except:
                        pass
                return None

            velocity = []
            for w in range(11, -1, -1):
                week_start = (
                    today - timedelta(days=today.weekday()) - timedelta(weeks=w)
                )
                week_end = week_start + timedelta(days=6)
                # Fetch all and filter in Python — DC dates are M/D/YYYY not ISO
                wrows = conn.execute(
                    f"""
                    SELECT b.m9_complete, b.m9_actual, b.m8_started, b.m8_actual
                    FROM blocked_data b JOIN accounts a ON a.account_id=b.account_id
                    WHERE 1=1 {_cohort_sql()}
                      AND (? = '' OR UPPER(COALESCE(b.account_theatre,a.account_theatre,'EMEA'))=UPPER(?))
                """,
                    (cohort, cohort, theatre, theatre),
                ).fetchall()
                n_m9 = sum(
                    1
                    for r in wrows
                    if r[0]
                    and _pdw(str(r[1] or ""))
                    and week_start <= _pdw(str(r[1])) <= week_end
                )
                n_m8 = sum(
                    1
                    for r in wrows
                    if r[2]
                    and _pdw(str(r[3] or ""))
                    and week_start <= _pdw(str(r[3])) <= week_end
                )
                velocity.append(
                    {
                        "week_start": str(week_start),
                        "week_end": str(week_end),
                        "m9_count": n_m9,
                        "m8_count": n_m8,
                        "label": week_start.strftime("%d %b"),
                    }
                )

        def parse_date(s):
            if not s:
                return None
            for f in ("%m/%d/%Y", "%Y-%m-%d", "%m/%d/%Y %H:%M:%S"):
                try:
                    from datetime import datetime as dt

                    return dt.strptime(s.strip().split(" ")[0], f).date()
                except:
                    pass
            return None

        next_targets = []
        overdue = []
        for r in targets:
            d = dict(r)
            m9d = parse_date(d.get("m9_planned"))
            if not m9d:
                continue
            subtype = d.get("subtype") or ""
            signal = d.get("signal") or ""
            m8 = bool(d["m8_started"])
            dc = d.get("dc_progress") or ""
            if subtype == "churn":
                confidence = "CHURN"
            elif m8 and dc == "Green" and signal != "blocked":
                confidence = "HIGH"
            elif m8 and (dc == "Yellow" or signal == "at_risk"):
                confidence = "MED"
            elif not m8 and dc == "Green" and signal == "green":
                confidence = "MED"
            else:
                confidence = "LOW"
            d["confidence"] = confidence
            d["m9_date"] = str(m9d)
            if m9d < today:
                d["status"] = "overdue"
                overdue.append(d)
            elif m9d <= three_month_end:
                d["status"] = "upcoming"
                next_targets.append(d)

        counts = [v["m9_count"] for v in velocity]
        trend = (
            "up"
            if counts[-1] > counts[0]
            else "down"
            if counts[-1] < counts[0]
            else "flat"
        )

        # Deduplicate by customer_name — same company can appear twice (case-variant IDs)
        seen_names = set()
        next_targets_dedup, overdue_dedup = [], []
        for r in next_targets:
            n = (r.get("customer_name") or "").lower().strip()
            if n not in seen_names:
                seen_names.add(n)
                next_targets_dedup.append(r)
        for r in overdue:
            n = (r.get("customer_name") or "").lower().strip()
            if n not in seen_names:
                seen_names.add(n)
                overdue_dedup.append(r)

        next_week_counts = {"EMEA": 0, "AMER": 0, "JAPAC": 0, "LATAM": 0}
        for r in next_targets_dedup:
            raw = r.get("m9_date") or r.get("m9_planned")
            m9d = parse_date(raw)
            if m9d and today <= m9d <= next_week_end:
                t = r.get("account_theatre") or "EMEA"
                if t in next_week_counts:
                    next_week_counts[t] += 1

        return {
            "next_targets": next_targets_dedup,
            "overdue": overdue_dedup,
            "velocity": velocity,
            "trend": trend,
            "next_week_counts": next_week_counts,
        }
    except Exception as e:
        return {"error": str(e)}


@app.get("/api/dq")
def api_dq():
    return _load_dq()


@app.get("/api/customer-search")
def api_customer_search(q: str = ""):
    """Search accounts by name — returns list of matches, deduplicated by name."""
    if not q or len(q) < 2:
        return []
    _ensure_db()
    try:
        with get_db() as conn:
            rows = conn.execute(
                """
                SELECT a.account_id, a.customer_name, a.active_cse, a.status,
                       a.sales_region, a.live_fire, b.signal, b.dc_progress
                FROM accounts a LEFT JOIN blocked_data b ON a.account_id=b.account_id
                WHERE a.customer_name LIKE ? AND a.customer_name!=''
                ORDER BY a.customer_name LIMIT 20
            """,
                (f"%{q}%",),
            ).fetchall()

        # Deduplicate by lower-cased customer_name: same company can have duplicate
        # account_ids that differ only in case (Salesforce ID normalisation issue).
        # Keep the row with the most DC data (signal + dc_progress present wins).
        # If both rows are genuinely distinct (different regions AND both have data),
        # keep both but mark them with a disambiguating region suffix in the name.
        seen: dict = {}  # lower(name) → list of row dicts
        for r in rows:
            d = dict(r)
            key = (d.get("customer_name") or "").lower().strip()
            seen.setdefault(key, []).append(d)

        result = []
        for key, candidates in seen.items():
            if len(candidates) == 1:
                result.append(candidates[0])
                continue
            # Multiple rows for same name — check if they're case-duplicates of the
            # same Salesforce ID or genuinely different accounts.
            ids_normalised = [c["account_id"].lower() for c in candidates]
            if len(set(ids_normalised)) == 1:
                # Pure case-duplicate: pick the one with more DC data
                def _score(c):
                    return (1 if c.get("signal") else 0) + (
                        1 if c.get("dc_progress") else 0
                    )

                best = max(candidates, key=_score)
                result.append(best)
            else:
                # Genuinely distinct accounts — keep all but make region visible
                # so the user can tell them apart in the dropdown.
                for c in candidates:
                    result.append(c)

        result.sort(key=lambda x: (x.get("customer_name") or "").lower())
        return result[:10]
    except:
        return []


@app.get("/api/customer/{account_id}")
def api_customer_detail(account_id: str):
    """Full customer card data."""
    _ensure_db()
    try:
        with get_db() as conn:
            r = conn.execute(
                """
                SELECT a.*, b.m0_complete, b.m1_complete, b.m1_planned,
                       b.m2_complete, b.m2_planned, b.m3_complete, b.m3_planned,
                       b.m4_complete, b.m4_planned, b.m5_complete, b.m5_planned,
                       b.m7_complete, b.m7_planned, b.m8_started, b.m8_planned, b.m8_actual,
                       b.m9_complete, b.m9_planned, b.m9_actual,
                       b.signal, b.subtype, b.dc_progress, b.status_detail,
                       b.upgrade_notes, b.health_notes, b.owner_e2e, b.dc_assignment,
                       b.cohort, b.area, b.region as district_region, b.district,
                       b.milestone_category, b.is_cs_team, b.team, b.exec_delay,
                       b.cc_rep, b.cc_dsm, b.churn_risk,
                       b.last_edited_by, b.last_edited_date, b.roadmap_url, b.ps_plan_url,
                       b.account_region, b.current_project_status, b.next_renewal_date,
                       b.past_due_planned, b.upgrade_duration_weeks,
                       b.has_partner, b.upgrade_partner,
                       b.m1_details, b.m3_details, b.m5_details,
                       b.milestone_aging, b.days_since_milestone, b.momentum_x,
                       b.entitlement_provision, b.activation_status, b.posture_workloads,
                       p.psc, p.pm, p.psc_shadow, p.ps_status, p.clarizen_id, p.timeline as ps_timeline,
                       ae.blocker, ae.owner, ae.accountable
                FROM accounts a
                LEFT JOIN blocked_data b ON a.account_id=b.account_id
                LEFT JOIN ps_data p ON a.account_id=p.account_id
                LEFT JOIN ai_enrichment ae ON a.account_id=ae.account_id
                WHERE a.account_id=?
            """,
                (account_id,),
            ).fetchone()
            # Status history
            hist = conn.execute(
                """
                SELECT old_status, new_status, changed_at, file_source, field_name
                FROM status_history WHERE account_id=? ORDER BY changed_at ASC LIMIT 20
            """,
                (account_id,),
            ).fetchall()
        if not r:
            raise HTTPException(status_code=404, detail="not found")
        d = dict(r)
        d["history"] = [dict(h) for h in hist]
        return d
    except HTTPException:
        raise
    except Exception as e:
        return {"error": str(e)}


def _generate_m1_rationale(cat: str, accounts: list, total: int) -> str:
    """Rationale built from actual upgrade_notes of accounts in this category."""
    seen = []
    for a in accounts:
        note = (a.get("notes") or "").strip()
        if note and note not in seen and note not in ("TBD", "-"):
            seen.append(note)
    return " | ".join(seen[:5]) if seen else ""


@app.get("/api/m0-needed")
def api_m0_needed(theatre: str = ""):
    """Accounts where M0 kickoff has not happened yet."""
    _ensure_db()
    try:
        with get_db() as conn:
            rows = conn.execute(
                """
                SELECT a.customer_name, a.active_cse, a.sales_region,
                       COALESCE(b.account_theatre, a.account_theatre, 'EMEA') as account_theatre,
                       b.signal, b.subtype, b.status_detail, b.health_notes,
                       b.dc_progress, b.cc_rep, b.cc_dsm, b.m1_planned,
                       a.account_id
                FROM accounts a JOIN blocked_data b ON a.account_id = b.account_id
                WHERE a.customer_name != ''
                  AND (b.m0_complete IS NULL OR b.m0_complete = 0)
                  AND b.m9_complete = 0
                  AND (? = '' OR UPPER(COALESCE(b.account_theatre, a.account_theatre, 'EMEA')) = UPPER(?))
                ORDER BY b.signal, a.sales_region, a.customer_name
                """,
                (theatre, theatre),
            ).fetchall()
        return [
            {
                "account_id": r["account_id"] or "",
                "customer_name": r["customer_name"] or "",
                "active_cse": r["active_cse"] or "",
                "sales_region": r["sales_region"] or "",
                "account_theatre": r["account_theatre"] or "",
                "signal": r["signal"] or "",
                "subtype": "m0_needed",
                "status_detail": r["status_detail"] or "",
                "health_notes": r["health_notes"] or "",
                "dc_progress": r["dc_progress"] or "",
                "cc_rep": r["cc_rep"] or "",
                "cc_dsm": r["cc_dsm"] or "",
                "m1_planned": r["m1_planned"] or "",
            }
            for r in rows
        ]
    except Exception:
        return []


@app.get("/api/m0-no-m1")
def api_m0_no_m1(theatre: str = ""):
    """Accounts where M0 kickoff done but M1 action plan not created yet."""
    _ensure_db()
    try:
        with get_db() as conn:
            rows = conn.execute(
                """
                SELECT a.customer_name, a.active_cse, a.sales_region,
                       COALESCE(b.account_theatre, a.account_theatre, 'EMEA') as account_theatre,
                       b.signal, b.subtype, b.status_detail, b.health_notes,
                       b.dc_progress, b.cc_rep, b.cc_dsm, b.m1_details, b.m1_planned,
                       a.account_id
                FROM accounts a JOIN blocked_data b ON a.account_id = b.account_id
                WHERE a.customer_name != ''
                  AND b.m0_complete = 1
                  AND (b.m1_complete IS NULL OR b.m1_complete = 0)
                  AND b.m9_complete = 0
                  AND (? = '' OR UPPER(COALESCE(b.account_theatre, a.account_theatre, 'EMEA')) = UPPER(?))
                ORDER BY b.signal, a.sales_region, a.customer_name
                """,
                (theatre, theatre),
            ).fetchall()
        return [
            {
                "account_id": r["account_id"] or "",
                "customer_name": r["customer_name"] or "",
                "active_cse": r["active_cse"] or "",
                "sales_region": r["sales_region"] or "",
                "account_theatre": r["account_theatre"] or "",
                "signal": r["signal"] or "",
                "subtype": "m0_no_m1",
                "status_detail": r["m1_details"] or r["status_detail"] or "",
                "health_notes": r["health_notes"] or "",
                "dc_progress": r["dc_progress"] or "",
                "cc_rep": r["cc_rep"] or "",
                "cc_dsm": r["cc_dsm"] or "",
                "m1_planned": r["m1_planned"] or "",
            }
            for r in rows
        ]
    except Exception:
        return []


@app.get("/api/m1-suggestions")
def api_m1_suggestions(theatre: str = ""):
    """M1 action plan — 4 flat tables with LLM rationale per category."""
    try:
        with get_db() as conn:
            rows = conn.execute(
                """
                SELECT s.assigned_cse, s.account_name, s.original_cse, s.region,
                       s.status, s.signal, s.category,
                       b.status_detail, b.upgrade_notes, b.health_notes,
                       b.cc_rep, b.cc_dsm, b.churn_risk, b.dc_progress,
                       a.live_fire, b.m3_planned, b.m8_planned, b.m9_planned,
                       a.account_id
                FROM m1_suggestions s
                LEFT JOIN accounts a ON a.customer_name=s.account_name
                LEFT JOIN blocked_data b ON b.account_id=a.account_id
                WHERE (? = '' OR UPPER(COALESCE(b.account_theatre,a.account_theatre,'EMEA'))=UPPER(?))
                ORDER BY s.category, s.assigned_cse, s.region, s.account_name
            """,
                (theatre, theatre),
            ).fetchall()
        from collections import defaultdict

        cats: dict = {"actionable": [], "acct_team": [], "unblock": [], "skip": []}
        for r in rows:
            cat = r[6] or "actionable"
            if cat in cats:
                # Pick best note: upgrade_notes first, then health_notes, then status_detail
                _EMPTY = {"tbd", "tbd.", "-", "—", "n/a", "none", ""}

                def _usable(s):
                    return s.strip() if s and s.strip().lower() not in _EMPTY else ""

                detail = _usable(r[7])  # status_detail
                notes = _usable(r[8])  # upgrade_notes
                health = _usable(r[9])  # health_notes
                best_note = notes or health or detail
                cats[cat].append(
                    {
                        "account_name": r[1],
                        "assigned_cse": r[0],
                        "original_cse": r[2],
                        "region": r[3],
                        "status": r[4],
                        "signal": r[5],
                        "is_new": r[2] == "Irene Garcia",
                        "status_detail": detail,
                        "notes": best_note[:200] if best_note else "",
                        "cc_rep": r[10] or "",
                        "cc_dsm": r[11] or "",
                        "churn_risk": r[12] or "",
                        "dc_progress": r[13] or "",
                        "live_fire": bool(r[14]),
                        "m3_planned": r[15] or "",
                        "m8_planned": r[16] or "",
                        "m9_planned": r[17] or "",
                    }
                )
        result = []
        for cat, label, color in [
            ("actionable", "✅ Ping Now", "green"),
            ("acct_team", "📞 Acct Team First", "amber"),
            ("unblock", "⚡ Single Unblock (MTN)", "blue"),
            ("skip", "🔴 Skip / Hard Blocked", "red"),
        ]:
            accs = cats[cat]
            rationale = _generate_m1_rationale(cat, accs, len(accs))
            result.append(
                {
                    "category": cat,
                    "label": label,
                    "color": color,
                    "total": len(accs),
                    "rationale": rationale,
                    "accounts": accs,
                }
            )
        return result
    except:
        return []


@app.get("/api/events")
async def api_events(request: Request):
    async def gen():
        while True:
            if await request.is_disconnected():
                break
            yield f"data: {json.dumps(_load_stats())}\n\n"
            await asyncio.sleep(30)

    return StreamingResponse(gen(), media_type="text/event-stream")


@app.get("/api/xsup-data")
def api_xsup_data(theatre: str = "", priority: str = "", cohort: str = ""):
    """Per-account XSUP summary — open XSUPs grouped by account, filterable by theatre/priority."""
    _ensure_db()
    try:
        with get_db() as conn:
            rows = conn.execute(
                """
                SELECT x.account_name, x.account_id, x.case_theatre,
                       x.xsup_number, x.xsup_priority, x.xsup_status,
                       x.case_status, x.summary, x.component, x.notes,
                       a.active_cse, a.sales_region,
                       COALESCE(a.account_theatre, x.case_theatre, '') as account_theatre
                FROM xsup_data x
                LEFT JOIN accounts a ON a.account_id = x.account_id
                WHERE (? = '' OR UPPER(COALESCE(a.account_theatre, x.case_theatre, ''))=UPPER(?))
                  AND (? = '' OR x.xsup_priority = ?)
                ORDER BY x.xsup_priority, x.account_name
                """,
                (theatre, theatre, priority, priority),
            ).fetchall()

        # Group by account
        from collections import defaultdict

        by_account: dict = {}
        for r in rows:
            d = dict(r)
            key = d["account_name"]
            if key not in by_account:
                by_account[key] = {
                    "account_name": key,
                    "account_id": d["account_id"],
                    "active_cse": d["active_cse"] or "",
                    "sales_region": d["sales_region"] or "",
                    "account_theatre": d["account_theatre"] or "",
                    "xsups": [],
                    "p1": 0,
                    "p2": 0,
                    "p3": 0,
                    "p4": 0,
                }
            by_account[key]["xsups"].append(
                {
                    "xsup_number": d["xsup_number"],
                    "xsup_priority": d["xsup_priority"],
                    "xsup_status": d["xsup_status"],
                    "case_status": d["case_status"],
                    "summary": d["summary"],
                    "component": d["component"],
                    "notes": d["notes"],
                }
            )
            p = (d["xsup_priority"] or "").lower()
            if p in ("p1", "p2", "p3", "p4"):
                by_account[key][p] += 1

        # Build set of XSUP numbers referenced in tech_blocker notes
        import re as _re

        _xpat = _re.compile(r"XSUP-\d+", _re.IGNORECASE)
        with get_db() as conn2:
            _tech_notes = conn2.execute(
                f"""
                SELECT b.upgrade_notes, b.health_notes, b.status_detail,
                       b.m1_details, b.m3_details, b.m5_details
                FROM blocked_data b
                WHERE b.subtype = 'tech_blocker'
                  {_cohort_sql()}
                """,
                (cohort, cohort),
            ).fetchall()
        _tech_xsup_refs: set = set()
        for _tn in _tech_notes:
            _combined = " ".join(filter(None, [_tn[i] or "" for i in range(6)]))
            _tech_xsup_refs.update(_xpat.findall(_combined))

        # Tag each account if any of its XSUPs appear in tech_blocker notes
        for acc in by_account.values():
            acc["also_tech_blocker"] = any(
                x["xsup_number"] in _tech_xsup_refs for x in acc["xsups"]
            )
            acc["tech_blocker_xsups"] = [
                x["xsup_number"]
                for x in acc["xsups"]
                if x["xsup_number"] in _tech_xsup_refs
            ]

        accounts = sorted(
            by_account.values(),
            key=lambda x: (-x["p1"], -x["p2"], -len(x["xsups"])),
        )
        total = sum(len(a["xsups"]) for a in accounts)
        return {"accounts": accounts, "total": total, "synced_at": _xsup_synced_at()}
    except Exception as e:
        return {"error": str(e), "accounts": [], "total": 0}


def _xsup_synced_at() -> str:
    try:
        with get_db() as conn:
            r = conn.execute(
                "SELECT synced_at FROM xsup_data ORDER BY id DESC LIMIT 1"
            ).fetchone()
            return r["synced_at"] if r else ""
    except Exception:
        return ""


@app.get("/api/sotu")
def api_sotu(theatre: str = "", cohort: str = ""):
    """State of the Union exec dashboard data."""
    from datetime import datetime as _dt

    _ensure_db()
    THEATRES = ["AMER", "EMEA", "JAPAC", "LATAM"]
    SUBTYPE_LABELS = {
        "customer_delay": "Customer Delay",
        "tech_blocker": "Tech Blocker",
        "core_rep_blocking": "Core Rep Blocking",
        "no_contact": "No Contact",
        "active_deal": "Active Deal",
        "legal_blocker": "Legal Blocker",
    }
    th_filter = theatre.upper() if theatre else ""

    with get_db() as conn:
        # ── Base filter matching Dashboard EXACTLY ───────────────────
        # NOT SH Only, NOT CC NNL, NOT Churn, NOT Excluded from SPO,
        # contract > 2025-07-31, cohort filter, theatre filter
        th_cond = "AND UPPER(COALESCE(b.account_theatre,'')) = ?" if th_filter else ""
        th_params = [th_filter] if th_filter else []

        BASE_FILTER = f"""
            b.pc_saas_vs_sh != 'SH Only'
            AND b.pc_cc_migration_status NOT IN ('CC NNL', 'Churn')
            AND b.excluded_from_spo != 'Excluded'
            AND b.last_contract_end_date > '2025-07-31'
            {_cohort_sql()}
            {th_cond}
        """

        def _count(extra: str = "", params: list = None) -> int:
            p = params or []
            sql = f"""SELECT COUNT(*) FROM blocked_data b
                      JOIN accounts a ON a.account_id = b.account_id
                      WHERE {BASE_FILTER}
                      {"AND " + extra if extra else ""}"""
            return conn.execute(sql, [cohort, cohort] + th_params + p).fetchone()[0]

        # Active SaaS = base filter (matches Dashboard Row 8)
        active_saas   = _count()
        # Indicated churn = field indicated churn within Active SaaS (Dashboard Row 9)
        indicated_churn = _count("b.field_indicated_churn = 'Sales indicated churn'")
        # To upgrade = Active - indicated churn (Dashboard Row 10 = 1325)
        in_scope      = active_saas - indicated_churn

        # Milestones within "to upgrade" population (excl field indicated churn)
        BASE_TO_UPGRADE = BASE_FILTER + " AND b.field_indicated_churn != 'Sales indicated churn'"

        def _ms_count(extra: str = "", params: list = None) -> int:
            p = params or []
            sql = f"""SELECT COUNT(*) FROM blocked_data b
                      JOIN accounts a ON a.account_id = b.account_id
                      WHERE {BASE_TO_UPGRADE}
                      {"AND " + extra if extra else ""}"""
            return conn.execute(sql, [cohort, cohort] + th_params + p).fetchone()[0]

        m9_complete   = _ms_count("b.m9_complete = 1")
        beat_plan     = _ms_count("b.m9_complete = 1 AND b.m9_planned != '' AND b.m9_actual < b.m9_planned")
        m8_inflight   = _ms_count("b.m8_started = 1 AND b.m9_complete = 0")
        pre_m8        = _ms_count("b.m8_started = 0 AND b.m9_complete = 0")

        # Stuck breakdown within pre-M8 (for stuck reason cards)
        stuck_total   = _ms_count("b.m8_started = 0 AND b.m9_complete = 0 AND b.subtype != ''")
        progressing   = _ms_count("b.m8_started = 0 AND b.m9_complete = 0 AND b.subtype = ''")
        churn         = indicated_churn  # rename for template compatibility

        # Confirmed churn since FY26 start (PC_CC_Migration_status='Churn', contract > 2025-07-31)
        # Excludes SH Only, NNL — these are accounts the business has written off
        confirmed_churn_sql = f"""
            SELECT COUNT(*) FROM blocked_data b
            JOIN accounts a ON a.account_id = b.account_id
            WHERE b.pc_cc_migration_status = 'Churn'
              AND b.pc_saas_vs_sh != 'SH Only'
              AND b.last_contract_end_date > '2025-07-31'
              {_cohort_sql()}
              {th_cond}
        """
        confirmed_churn = conn.execute(confirmed_churn_sql, [cohort, cohort] + th_params).fetchone()[0]

        # ── Stuck reasons (pre-M8, to-upgrade population) ───────────
        stuck_sql = f"""
            SELECT b.subtype, b.account_theatre, COUNT(*) as cnt
            FROM blocked_data b JOIN accounts a ON a.account_id = b.account_id
            WHERE {BASE_TO_UPGRADE}
              AND b.m9_complete = 0
              AND b.m8_started = 0
              AND b.subtype != ''
            GROUP BY b.subtype, b.account_theatre
            ORDER BY b.subtype, b.account_theatre
        """
        stuck_rows = conn.execute(stuck_sql, [cohort, cohort] + th_params).fetchall()

        stuck_by_type: dict = {}
        for subtype, theatre_val, cnt in stuck_rows:
            if subtype not in stuck_by_type:
                stuck_by_type[subtype] = {
                    "subtype": subtype,
                    "label": SUBTYPE_LABELS.get(subtype, subtype),
                    "total": 0,
                    "by_theatre": {},
                }
            stuck_by_type[subtype]["total"] += cnt
            stuck_by_type[subtype]["by_theatre"][theatre_val or "Unknown"] = cnt

        subtype_order = [
            "customer_delay",
            "tech_blocker",
            "core_rep_blocking",
            "no_contact",
            "active_deal",
            "legal_blocker",
        ]
        stuck = [stuck_by_type[k] for k in subtype_order if k in stuck_by_type]
        # Append any unexpected subtypes at end
        for k, v in stuck_by_type.items():
            if k not in subtype_order:
                stuck.append(v)

        # ── Historical completions (from m9_actual in blocked_data) ─────
        hist_th_cond = (
            "AND UPPER(COALESCE(b.account_theatre,'')) = ?" if th_filter else ""
        )
        hist_sql = f"""
            SELECT strftime('%Y-%m', substr(b.m9_actual, 1, 10)) as month,
                   COALESCE(b.account_theatre, a.account_theatre, 'Unknown') as theatre,
                   COUNT(*) as cnt
            FROM blocked_data b
            JOIN accounts a ON a.account_id = b.account_id
            WHERE b.m9_complete = 1
              AND b.m9_actual IS NOT NULL AND b.m9_actual != ''
              AND substr(b.m9_actual, 1, 7) >= '2026-01'
              {_cohort_sql()}
              {hist_th_cond}
            GROUP BY month, theatre
            ORDER BY month, theatre
        """
        hist_rows = conn.execute(hist_sql, [cohort, cohort] + th_params).fetchall()

        comp_by_month: dict = {}
        for month, theatre_val, cnt in hist_rows:
            if month not in comp_by_month:
                comp_by_month[month] = {t: 0 for t in THEATRES}
                comp_by_month[month]["month"] = month
            if theatre_val in THEATRES:
                comp_by_month[month][theatre_val] = cnt

        completions = []
        for month in sorted(comp_by_month.keys()):
            row = comp_by_month[month]
            row["total"] = sum(row.get(t, 0) for t in THEATRES)
            completions.append(row)

        # ── Forecast (m9_planned, non-churn, non-complete) ───────────
        fcast_th_cond = (
            "AND UPPER(COALESCE(b.account_theatre,'')) = ?" if th_filter else ""
        )
        fcast_sql = f"""
            SELECT b.m9_planned, b.account_theatre
            FROM blocked_data b JOIN accounts a ON a.account_id = b.account_id
            WHERE 1=1 {_cohort_sql()}
              AND b.m9_complete = 0
              AND b.subtype != 'churn'
              AND b.m9_planned IS NOT NULL AND b.m9_planned != ''
              {fcast_th_cond}
        """
        fcast_rows = conn.execute(fcast_sql, [cohort, cohort] + th_params).fetchall()

        fcast_by_month: dict = {}
        for m9p, theatre_val in fcast_rows:
            try:
                _s = str(m9p).strip()[:10]  # works for both MM/DD/YYYY and YYYY-MM-DD HH:MM:SS
                for _fmt in ("%Y-%m-%d", "%m/%d/%Y"):
                    try:
                        d = _dt.strptime(_s, _fmt)
                        break
                    except ValueError:
                        continue
                else:
                    continue
                ym = d.strftime("%Y-%m")
                if ym < "2026-07":
                    continue
                if ym not in fcast_by_month:
                    fcast_by_month[ym] = {t: 0 for t in THEATRES}
                    fcast_by_month[ym]["month"] = ym
                if (theatre_val or "") in THEATRES:
                    fcast_by_month[ym][theatre_val] = (
                        fcast_by_month[ym].get(theatre_val, 0) + 1
                    )
            except (ValueError, TypeError):
                pass

        forecast = []
        for month in sorted(fcast_by_month.keys())[:8]:
            row = fcast_by_month[month]
            row["total"] = sum(row.get(t, 0) for t in THEATRES)
            forecast.append(row)

        # ── Adjusted forecast — all-time avg run rate (Jan–last full month) ──
        rate_sql = f"""
            SELECT COALESCE(b.account_theatre, a.account_theatre, 'Unknown') as theatre,
                   COUNT(*) as cnt
            FROM blocked_data b
            JOIN accounts a ON a.account_id = b.account_id
            WHERE b.m9_complete = 1
              AND b.m9_actual IS NOT NULL AND b.m9_actual != ''
              AND substr(b.m9_actual, 1, 7) >= '2026-01'
              AND substr(b.m9_actual, 1, 7) < strftime('%Y-%m', date('now', 'start of month'))
              {_cohort_sql()}
              {hist_th_cond}
            GROUP BY theatre
        """
        rate_rows = conn.execute(rate_sql, [cohort, cohort] + th_params).fetchall()
        rate_by_theatre = {r[0]: r[1] for r in rate_rows if r[0] in THEATRES}
        # Count completed full months
        n_rate_months = (
            conn.execute(f"""
            SELECT COUNT(DISTINCT strftime('%Y-%m', substr(b.m9_actual, 1, 10)))
            FROM blocked_data b
            JOIN accounts a ON a.account_id = b.account_id
            WHERE b.m9_complete = 1
              AND b.m9_actual IS NOT NULL AND b.m9_actual != ''
              AND substr(b.m9_actual, 1, 7) >= '2026-01'
              AND substr(b.m9_actual, 1, 7) < strftime('%Y-%m', date('now', 'start of month'))
              {_cohort_sql()}
        """, (cohort, cohort)).fetchone()[0]
            or 1
        )
        monthly_rate = {
            t: max(1, round(rate_by_theatre.get(t, 0) / n_rate_months))
            for t in THEATRES
        }
        total_rate = sum(monthly_rate.values())

        # FY27: Aug 2026 → Aug 2027 (13 months)
        # FY27: Aug 2026 → Jul 2027 (12 months)
        _fy27_months = [
            f"{y}-{m:02d}"
            for y, m in [
                (2026,8),(2026,9),(2026,10),(2026,11),(2026,12),
                (2027,1),(2027,2),(2027,3),(2027,4),(2027,5),(2027,6),(2027,7),
            ]
        ]
        # Pull any existing planned data; fill missing months with flat rate
        planned_by_month = {row["month"]: row for row in forecast}
        adjusted_forecast = []
        for ym in _fy27_months:
            base = planned_by_month.get(ym, {})
            adj = {t: monthly_rate[t] for t in THEATRES}
            adj["month"] = ym
            adj["total"] = total_rate
            adjusted_forecast.append(adj)

        # ── Monte Carlo — release-aware per-month simulation ─────────
        # Lessons learned:
        #   • Wave pattern: completions spike every ~2 months (observed Jan-Jun 2026)
        #   • Jul/Aug release → uptick lands Sep/Oct (6-8 wks later)
        #   • Nov/Dec release → uptick lands Jan/Feb
        #   • Feb release    → uptick lands Mar/Apr
        #   • 24 FR-blocked accounts waiting on releases (COE tracker)
        #   • 38 accounts with open XSUPs dragging pace by ~2/month
        import random as _random
        import statistics as _stats

        mc_hist_rows = conn.execute(f"""
            SELECT strftime('%Y-%m', substr(b.m9_actual, 1, 10)) as month, COUNT(*) as cnt
            FROM blocked_data b
            JOIN accounts a ON a.account_id = b.account_id
            WHERE b.m9_complete = 1
              AND b.m9_actual IS NOT NULL AND b.m9_actual != ''
              AND substr(b.m9_actual, 1, 7) >= '2026-01'
              AND substr(b.m9_actual, 1, 7) < strftime('%Y-%m', date('now', 'start of month'))
              {_cohort_sql()}
            GROUP BY month ORDER BY month
        """, (cohort, cohort)).fetchall()
        mc_sample = [r[1] for r in mc_hist_rows] if mc_hist_rows else [10]
        mc_mean = _stats.mean(mc_sample) if mc_sample else 10
        mc_std  = _stats.stdev(mc_sample) if len(mc_sample) > 1 else 4

        def _pct(data, p):
            s = sorted(data)
            k = (len(s) - 1) * p / 100
            lo, hi = int(k), min(int(k) + 1, len(s) - 1)
            return round(s[lo] + (s[hi] - s[lo]) * (k - lo))

        # Per-month expected rate = base rate adjusted for release timing
        # Quiet months: base - xsup_drag (~2/month)
        # Release uptick months: base + FR_unlock_pulse + wave_boost
        _RELEASE_UPLIFT = {
            # Sep/Oct = Jul release lands (6-8 wks later)
            "2026-09": mc_mean * 1.35,
            "2026-10": mc_mean * 1.50,
            # Dec/Jan = Nov release lands (6-8 wks later)
            "2026-12": mc_mean * 1.25,
            "2027-01": mc_mean * 1.40,
            # Mar/Apr = Feb release lands (6-8 wks later)
            "2027-03": mc_mean * 1.25,
            "2027-04": mc_mean * 1.40,
        }
        _XSUP_DRAG = 2.0  # 38 open-XSUP accounts slow throughput ~2/month

        def _month_rate(ym: str) -> float:
            if ym in _RELEASE_UPLIFT:
                return _RELEASE_UPLIFT[ym]
            return max(6, mc_mean - _XSUP_DRAG)

        _random.seed(42)
        n_sim = 10_000

        # Per-month bands: simulate each month with its own expected rate
        mc_per_month_by_month = {}
        for row in adjusted_forecast:
            ym = row["month"]
            rate = _month_rate(ym)
            col = sorted([
                max(0, round(_random.gauss(rate, mc_std * 0.6)))
                for _ in range(n_sim)
            ])
            mc_per_month_by_month[ym] = {
                "p10": col[int(n_sim * .10)],
                "p30": col[int(n_sim * .30)],
                "p50": col[int(n_sim * .50)],
                "p70": col[int(n_sim * .70)],
                "p90": col[int(n_sim * .90)],
                "rate": round(rate, 1),
            }

        # mc_per_month = first forecast month (for legend display)
        first_ym = adjusted_forecast[0]["month"] if adjusted_forecast else "2026-07"
        mc_per_month = mc_per_month_by_month.get(first_ym, {"p10":6,"p30":9,"p50":13,"p70":16,"p90":19})

        # FY27 end (Aug 2027): confirmed so far + all 13 months simulated
        confirmed_ytd = conn.execute(f"""
            SELECT COUNT(*) FROM blocked_data b
            JOIN accounts a ON a.account_id = b.account_id
            WHERE b.m9_complete = 1
              AND b.m9_actual IS NOT NULL AND b.m9_actual != ''
              AND substr(b.m9_actual, 1, 7) >= '2026-01'
              {_cohort_sql()}
        """, (cohort, cohort)).fetchone()[0]

        fy27_sims = []
        for _ in range(n_sim):
            total = confirmed_ytd
            for row in adjusted_forecast:
                rate = _month_rate(row["month"])
                total += max(0, round(_random.gauss(rate, mc_std * 0.6)))
            fy27_sims.append(total)

        mc_year_end = {
            "confirmed": confirmed_ytd,
            "label": "FY27 end (Jul 2027)",
            "p10": _pct(fy27_sims, 10),
            "p30": _pct(fy27_sims, 30),
            "p50": _pct(fy27_sims, 50),
            "p70": _pct(fy27_sims, 70),
            "p90": _pct(fy27_sims, 90),
        }
        mc_sample_mean = round(mc_mean, 1)
        mc_sample_n = len(mc_sample)

    return {
        "kpi": {
            "active_saas": active_saas,       # Dashboard Row 8
            "indicated_churn": indicated_churn, # Dashboard Row 9 (field-indicated only)
            "in_scope": in_scope,              # Dashboard Row 10 = to upgrade = 1325
            "m9_complete": m9_complete,        # Dashboard Row 20 = 98
            "beat_plan": beat_plan,
            "m8_inflight": m8_inflight,        # in-flight (started, not done)
            "pre_m8": pre_m8,                  # not yet in M8, not done
            "churn": churn,                    # = indicated_churn (sales flagged, still active)
            "confirmed_churn": confirmed_churn, # PC_CC_Migration_status=Churn, fully written off
            "stuck_total": stuck_total,
            "progressing": progressing,
        },
        "stuck": stuck,
        "completions": completions,
        "forecast": adjusted_forecast,
        "run_rate": {t: monthly_rate[t] for t in THEATRES},
        "run_rate_total": total_rate,
        "run_rate_months": n_rate_months,
        "monte_carlo": {
            "per_month": mc_per_month,            # first month (for legend)
            "by_month": mc_per_month_by_month,    # per-month bands with release uplift
            "year_end": mc_year_end,
            "sample_n": mc_sample_n,
            "sample_mean": mc_sample_mean,
        },
    }


@app.get("/api/velocity")
def api_velocity(weeks: int = 12, theatre: str = "", cohort: str = ""):
    """Milestone velocity — this week summary + N-week history by region."""
    from datetime import date as _date, timedelta, datetime as _dt, timezone

    _ensure_db()

    MILESTONES = [
        "M1 Outreach",
        "M2 Entitlements",
        "M3 Buy-in",
        "M4 Discovery",
        "M5 Tech Validation",
        "M8 Upgrade Started",
        "M9 Upgrade Complete",
    ]
    THEATRES = ["AMER", "EMEA", "JAPAC", "LATAM"]

    today = _dt.now(timezone.utc).date()
    this_monday = today - timedelta(days=today.weekday())

    weeks = max(1, min(weeks, 104))  # cap at 2 years

    def _week_label(monday: _date) -> str:
        return monday.strftime("%b") + " " + str(monday.day)

    def _count_week(monday: _date, theatre_filter: str) -> dict:
        """Count milestone completions in the given Mon–Sun window."""
        mon_s = monday.isoformat()
        next_monday_s = (monday + timedelta(days=7)).isoformat()
        t_clause = (
            "AND UPPER(COALESCE(b.account_theatre, a.account_theatre,'EMEA'))=UPPER(?)"
            if theatre_filter
            else ""
        )
        t_params = (theatre_filter,) if theatre_filter else ()
        result = {}
        with get_db() as conn:
            for ms in MILESTONES:
                row = conn.execute(
                    f"""
                    SELECT COUNT(DISTINCT sh.account_id) as cnt
                    FROM status_history sh
                    JOIN accounts a ON a.account_id = sh.account_id
                    LEFT JOIN blocked_data b ON b.account_id = sh.account_id
                    WHERE sh.field_name = ?
                      AND sh.new_status = 'Y'
                      AND sh.changed_at >= ?
                      AND sh.changed_at < ?
                      {_cohort_sql()}
                      {t_clause}
                    """,
                    (ms, mon_s, next_monday_s, cohort, cohort) + t_params,
                ).fetchone()
                cnt = row["cnt"] if row else 0
                if cnt:
                    result[ms] = cnt
        return result

    # This week — always all theatres
    this_week_by_theatre = {}
    for t in THEATRES:
        this_week_by_theatre[t] = _count_week(this_monday, t)

    sun = this_monday + timedelta(days=6)
    range_label = f"{this_monday.strftime('%b') + ' ' + str(this_monday.day)} – {sun.strftime('%b') + ' ' + str(sun.day)}"

    # History — N weeks ending last Sunday
    history = []
    last_monday = this_monday - timedelta(weeks=1)
    for i in range(weeks):
        w_monday = last_monday - timedelta(weeks=i)
        counts = _count_week(w_monday, theatre)
        row = {"week": _week_label(w_monday)}
        row.update(counts)
        history.append(row)

    return {
        "this_week": {
            "range": range_label,
            "by_theatre": this_week_by_theatre,
        },
        "history": history,
        "updated_at": _dt.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S"),
    }


@app.get("/velocity", response_class=HTMLResponse)
def page_velocity():
    html_path = Path(__file__).parent / "static" / "velocity.html"
    if html_path.exists():
        return html_path.read_text()
    return "<h1>velocity.html not found</h1>"


@app.get("/", response_class=HTMLResponse)
def dashboard():
    from fastapi.responses import RedirectResponse

    return RedirectResponse(url="/ops")


@app.get("/ops", response_class=HTMLResponse)
def dashboard_ops():
    return dashboard_v2()


@app.get("/v2", response_class=HTMLResponse)
def dashboard_v2():
    html_path = Path(__file__).parent / "static" / "v2.html"
    if html_path.exists():
        return html_path.read_text()
    return "<h1>v2 not found</h1>"


@app.get("/daily", response_class=HTMLResponse)
def dashboard_daily():
    html_path = Path(__file__).parent / "static" / "daily.html"
    if html_path.exists():
        return html_path.read_text()
    return "<h1>daily not found</h1>"


@app.get("/blockers", response_class=HTMLResponse)
def page_blockers():
    html_path = Path(__file__).parent / "static" / "blockers.html"
    if html_path.exists():
        return html_path.read_text()
    return "<h1>blockers page not found</h1>"


@app.get("/forecast", response_class=HTMLResponse)
def page_forecast():
    html_path = Path(__file__).parent / "static" / "forecast.html"
    if html_path.exists():
        return html_path.read_text()
    return "<h1>forecast page not found</h1>"


@app.get("/audit", response_class=HTMLResponse)
def page_audit():
    html_path = Path(__file__).parent / "static" / "audit.html"
    if html_path.exists():
        return html_path.read_text()
    return "<h1>audit page not found</h1>"


@app.get("/sotu", response_class=HTMLResponse)
def page_sotu():
    html_path = Path(__file__).parent / "static" / "sotu.html"
    if html_path.exists():
        return html_path.read_text()
    return "<h1>sotu.html not found</h1>"


@app.get("/cse", response_class=HTMLResponse)
def dashboard_cse():
    return (Path(__file__).parent / "static" / "cse.html").read_text(encoding="utf-8")


@app.get("/weekly", response_class=HTMLResponse)
def dashboard_weekly():
    return (Path(__file__).parent / "static" / "weekly.html").read_text(
        encoding="utf-8"
    )


@app.get("/pc-cc", response_class=HTMLResponse)
def dashboard_pc_cc():
    return (Path(__file__).parent / "static" / "pc_cc.html").read_text(encoding="utf-8")


@app.get("/api/pc-cc-accounts")
def api_pc_cc_accounts(cohort: str = ""):
    import csv as _csv

    data_dir = Path(__file__).parent / "data"
    accounts_file = data_dir / "pc_cc_accounts.csv"
    mapping_file = data_dir / "cc_resource_mapping.csv"
    if not accounts_file.exists() or not mapping_file.exists():
        return []

    def _norm(s):
        return s.strip().lower() if s else ""

    region_map = {}
    with open(mapping_file) as f:
        for r in _csv.DictReader(f):
            reg = _norm(r.get("Region", ""))
            if not reg:
                continue
            rsm = r.get("CC-RSM", "").strip()
            dsm = r.get("CC-DSM", "").strip()
            area = r.get("Area", "").strip()
            if (
                reg not in region_map
                or "no coverage" in region_map[reg]["cc_rsm"].lower()
            ):
                region_map[reg] = {"area": area, "cc_rsm": rsm, "cc_dsm": dsm}

    results = []
    with open(accounts_file) as f, get_db() as conn:
        for acc in _csv.DictReader(f):
            sid = acc.get("sfdc_account_id", "").strip().lower()
            name = acc.get("sfdc_account_name", "").strip()
            region = acc.get("account_region", "").strip()
            mig = acc.get("PC_CC_Migration_status", "").strip()

            db_row = conn.execute(
                f"SELECT a.active_cse FROM accounts a "
                f"JOIN blocked_data b ON a.account_id=b.account_id "
                f"WHERE LOWER(a.account_id)=? {_cohort_sql()}",
                (sid, cohort, cohort),
            ).fetchone()

            rm = region_map.get(_norm(region), {})
            cc_rsm = rm.get("cc_rsm", "No Coverage")
            cc_dsm = rm.get("cc_dsm", "")
            area = rm.get("area", "")
            cse = (
                db_row["active_cse"] if db_row and db_row["active_cse"] else "— No CSE"
            )
            no_rsm = "no coverage" in cc_rsm.lower() or "no match" in cc_rsm.lower()

            results.append(
                {
                    "account_id": sid,
                    "name": name,
                    "cse": cse,
                    "area": area,
                    "region": region,
                    "migration": mig,
                    "priority": "PC to CC Migration" in mig,
                    "cc_rsm": cc_rsm,
                    "cc_dsm": cc_dsm,
                    "no_rsm": no_rsm,
                    "in_solstice": bool(db_row),
                }
            )

    results.sort(key=lambda x: (x["cse"], x["name"]))
    return results


@app.get("/scope", response_class=HTMLResponse)
def dashboard_scope():
    return (Path(__file__).parent / "static" / "scope.html").read_text(encoding="utf-8")


@app.get("/api/scope")
def api_scope(theatre: str = "", cohort: str = ""):
    """In-scope accounts — not churned, M9 not complete. Empty theatre/cohort = all."""
    _ensure_db()
    try:
        with get_db() as conn:
            rows = conn.execute(
                f"""
                SELECT b.account_id,
                       a.customer_name, a.active_cse, a.sales_region, a.status,
                       b.account_theatre, b.cohort, b.area, b.dc_progress, b.signal, b.subtype,
                       b.status_detail, b.upgrade_notes, b.health_notes, b.notes,
                       b.cc_rep, b.cc_dsm, b.churn_risk,
                       b.m0_complete, b.m1_complete, b.m2_complete, b.m3_complete,
                       b.m4_complete, b.m5_complete, b.m6_complete, b.m7_complete,
                       b.m8_started, b.m9_complete,
                       b.m8_planned, b.m9_planned, b.m8_actual, b.m9_actual,
                       b.m1_details, b.m3_details, b.m5_details,
                       b.milestone_category, b.has_partner, b.upgrade_partner,
                       b.next_renewal_date, b.current_project_status
                FROM blocked_data b
                JOIN accounts a ON a.account_id = b.account_id
                WHERE 1=1 {_cohort_sql()}
                  AND (? = '' OR UPPER(b.account_theatre) = UPPER(?))
                  AND (a.status IS NULL OR a.status != 'Churning/Churned')
                  AND (b.status_detail IS NULL OR b.status_detail NOT LIKE '%decided to churn%')
                  AND (b.m9_complete IS NULL OR b.m9_complete != 1)
                ORDER BY b.account_theatre, a.active_cse, a.customer_name
                """,
                (cohort, cohort, theatre, theatre),
            ).fetchall()
        return [dict(r) for r in rows]
    except Exception as e:
        logger.error("api_scope failed: %s", e)
        return []


@app.get("/compare", response_class=HTMLResponse)
def dashboard_compare():
    from fastapi.responses import RedirectResponse

    return RedirectResponse(url="/ops", status_code=301)


@app.get("/api/daily-brief")
def api_daily_brief(date: str = "", theatre: str = "", cohort: str = ""):
    """Leadership daily briefing — movements for a given date + 7-day trend."""
    _ensure_db()
    from datetime import datetime, timedelta

    if not date:
        from datetime import date as _date

        date = _date.today().isoformat()
    try:
        target = datetime.strptime(date, "%Y-%m-%d")
    except:
        return {"error": "Invalid date"}
    try:
        with get_db() as conn:
            # All movements for this day
            day_rows = conn.execute(
                f"""
                SELECT sh.account_id, sh.field_name, sh.old_status, sh.new_status,
                       sh.changed_at, sh.file_source,
                       CASE WHEN sh.account_id='unmatched_dc' THEN sh.new_status
                            ELSE a.customer_name END as customer_name,
                       CASE WHEN sh.account_id='unmatched_dc' THEN sh.old_status
                            ELSE COALESCE(a.active_cse,'—') END as active_cse,
                       COALESCE(b.area,'—') as area,
                       COALESCE(a.sales_region,'—') as region,
                       COALESCE(b.churn_risk,'') as churn_risk
                FROM status_history sh
                LEFT JOIN accounts a ON a.account_id=sh.account_id
                LEFT JOIN blocked_data b ON b.account_id=sh.account_id
                WHERE sh.changed_at >= ? AND sh.changed_at < ?
                  AND sh.source IN ('pipeline','backfill')
                  AND (sh.account_id='unmatched_dc' OR {_cohort_sql()})
                  AND (? = ''
                       OR UPPER(COALESCE(a.account_theatre,'EMEA'))=UPPER(?)
                       OR (sh.account_id='unmatched_dc' AND (
                           ? = '' OR UPPER(COALESCE(
                             (SELECT a2.account_theatre FROM accounts a2
                              WHERE a2.customer_name=sh.new_status LIMIT 1),
                             'EMEA'))=UPPER(?))))
                ORDER BY sh.changed_at
            """,
                (
                    date + "T00:00:00",
                    date + "T23:59:59",
                    cohort,
                    cohort,
                    theatre,
                    theatre,
                    theatre,
                    theatre,
                ),
            ).fetchall()

            # 7-day trend
            trend = []
            for i in range(29, -1, -1):
                d = (target - timedelta(days=i)).strftime("%Y-%m-%d")
                r = conn.execute(
                    """
                    SELECT
                      SUM(CASE WHEN sh.field_name='M8 Upgrade Started' THEN 1 ELSE 0 END) m8,
                      SUM(CASE WHEN sh.field_name='M9 Upgrade Complete' THEN 1 ELSE 0 END) m9,
                      0 m3, 0 m1,
                      SUM(CASE WHEN sh.field_name IN ('M8 Upgrade Started','M9 Upgrade Complete') THEN 1 ELSE 0 END) total
                    FROM status_history sh
                    LEFT JOIN accounts a ON a.account_id=sh.account_id
                    WHERE sh.changed_at >= ? AND sh.changed_at < ?
                      AND sh.source IN ('pipeline','backfill')
                      AND sh.field_name IN ('M8 Upgrade Started','M9 Upgrade Complete')
                      AND sh.new_status='Y'
                      AND (? = ''
                           OR UPPER(COALESCE(a.account_theatre,'EMEA'))=UPPER(?)
                           OR (sh.account_id='unmatched_dc' AND (
                               ? = '' OR UPPER(COALESCE(
                                 (SELECT a2.account_theatre FROM accounts a2
                                  WHERE a2.customer_name=sh.new_status LIMIT 1),
                                 'EMEA'))=UPPER(?))))
                """,
                    (
                        d + "T00:00:00",
                        d + "T23:59:59",
                        theatre,
                        theatre,
                        theatre,
                        theatre,
                    ),
                ).fetchone()
                trend.append(
                    {
                        "date": d,
                        "m8": r[0] or 0,
                        "m9": r[1] or 0,
                        "m3": r[2] or 0,
                        "m1": r[3] or 0,
                        "total": r[4] or 0,
                    }
                )

            # Per-theatre cumulative breakdown
            theatre_totals = {}
            for t in ["EMEA", "JAPAC", "AMER", "LATAM"]:
                tr = conn.execute(
                    f"""
                    SELECT SUM(b.m8_started) m8, SUM(b.m9_complete) m9,
                           SUM(b.m3_complete) m3, COUNT(*) accounts
                    FROM blocked_data b JOIN accounts a ON a.account_id=b.account_id
                    WHERE UPPER(COALESCE(a.account_theatre,'EMEA'))=?
                      {_cohort_sql()}
                """,
                    (t, cohort, cohort),
                ).fetchone()
                theatre_totals[t] = {
                    "m8": tr[0] or 0,
                    "m9": tr[1] or 0,
                    "m3": tr[2] or 0,
                    "accounts": tr[3] or 0,
                }

            # Cumulative totals (filtered or global)
            totals = conn.execute(
                f"""
                SELECT
                  SUM(b.m8_started) m8, SUM(b.m9_complete) m9,
                  SUM(b.m3_complete) m3, SUM(b.m5_complete) m5,
                  COUNT(*) accounts
                FROM blocked_data b JOIN accounts a ON a.account_id=b.account_id
                WHERE 1=1 {_cohort_sql()}
                  AND (? = '' OR UPPER(COALESCE(a.account_theatre,'EMEA'))=UPPER(?))
            """,
                (cohort, cohort, theatre, theatre),
            ).fetchone()

        movements = [dict(r) for r in day_rows]
        # Headline counts
        headline = {
            "m8_started": sum(
                1 for m in movements if m["field_name"] == "M8 Upgrade Started"
            ),
            "m9_complete": sum(
                1 for m in movements if m["field_name"] == "M9 Upgrade Complete"
            ),
            "m3_complete": sum(1 for m in movements if m["field_name"] == "M3 Buy-in"),
            "m1_outreach": sum(
                1 for m in movements if m["field_name"] == "M1 Outreach"
            ),
            "cse_changes": sum(1 for m in movements if m["field_name"] == "cse"),
            "status_changes": sum(1 for m in movements if m["field_name"] == "status"),
            "regressions": sum(
                1
                for m in movements
                if m.get("new_status") in ("N", "") and m.get("old_status") == "Y"
            ),
            "total": len(movements),
        }
        return {
            "date": date,
            "theatre": theatre or "All",
            "headline": headline,
            "movements": movements,
            "trend": trend,
            "theatre_totals": theatre_totals,
            "cumulative": {
                "m8_total": totals["m8"] or 0,
                "m9_total": totals["m9"] or 0,
                "m3_total": totals["m3"] or 0,
                "m5_total": totals["m5"] or 0,
                "accounts": totals["accounts"] or 0,
            },
        }
    except Exception as e:
        return {"error": str(e)}


if __name__ == "__main__":
    import os as _os

    # Always run from Solstice/ directory regardless of where called from
    _os.chdir(Path(__file__).parent)
    sys.path.insert(0, str(Path(__file__).parent))

    import json as _j
    from agent.ps_parser import load_and_merge as _mp
    from agent.db import migrate_from_state as _mig

    # Always populate DB from all CSVs on startup
    init_db()
    print("Loading data sources...")

    if (DATA_DIR / "ps_tracker.csv").exists():
        _mp(DATA_DIR / "ps_tracker.csv", STATE_FILE)
        print("  ✓ PS tracker merged")
    if (DATA_DIR / "unified_tracker2.xlsx").exists():
        try:
            _run_dc_pipeline(DATA_DIR, STATE_FILE)
            print(
                "  ✓ Unified Tracker 2.0 synced via _run_dc_pipeline (all theatres, milestones)"
            )
        except Exception as _dc_err:
            print(
                f"  ⚠ Unified Tracker 2.0 skipped — parse error: {_dc_err}"
            )
    try:
        _mig(STATE_FILE)
    except Exception as _mig_err:
        print(f"  ⚠ state migration skipped: {_mig_err}")
    # Sync live_fire into DB
    _state = _j.loads(STATE_FILE.read_text())
    with get_db() as _conn:
        for _aid, _acc in _state.get("accounts", {}).items():
            _conn.execute(
                "UPDATE accounts SET live_fire=?, live_fire_dc=? WHERE account_id=?",
                (
                    1 if _acc.get("live_fire") else 0,
                    _acc.get("live_fire_dc", "") or "",
                    _aid,
                ),
            )
    with get_db() as _conn:
        _n = _conn.execute("SELECT COUNT(*) FROM blocked_data").fetchone()[0]
        _lf = _conn.execute(
            "SELECT COUNT(*) FROM accounts WHERE live_fire=1"
        ).fetchone()[0]
    print(f"  ✓ DB ready: {_n} milestone records | {_lf} live fire accounts")
    print("Solstice Control Center → http://localhost:8200")
    uvicorn.run(app, host="0.0.0.0", port=8200, log_level="warning")
