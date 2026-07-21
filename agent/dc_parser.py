"""
dc_parser.py — Parser for the DC CSE Tracker (master source of truth).

This is the authoritative file. Its data wins over all other sources on conflict.
Filters to EMEA accounts only. Merges into state.json by account_id (Salesforce ID).

Key fields extracted:
  - CSE Assigned       → active_cse  (wins over EMEA accounts CSV)
  - PC_CC_Migration_status → dc_status
  - All M0-M9 milestone dates and completion flags
  - Email sent         → email_sent
  - Status Detail      → status_detail
  - DC assignment      → dc_assignment
  - Owner: End to end upgrade → owner_e2e
"""

from __future__ import annotations
import csv
import json
import logging
import re
from datetime import datetime, timezone
from pathlib import Path

logger = logging.getLogger(__name__)

BOOL_YES = {"y", "yes", "true", "1", "x"}


def _yn(val: str) -> bool:
    return val.strip().lower() in BOOL_YES


def _signal_from_detail(status_detail: str) -> str:
    """Derive signal from DC Status Detail emoji prefix. Returns '' when no emoji recognised."""
    d = status_detail.strip()
    if d.startswith("\U00002705") or d.startswith("✅"):
        return "green"
    if d.startswith("\U0001f6d1") or d.startswith("🛑"):
        return "blocked"
    if d.startswith("\U0001f44e") or d.startswith("👎"):
        return "at_risk"
    return ""  # empty status_detail or unrecognised prefix — not 'at_risk' by default


def _subtype_from_detail(status_detail: str) -> str:
    """Derive blocker subtype from DC Status Detail text."""
    d = status_detail.lower()
    # Churn / lost — highest priority, surface immediately
    if (
        "decided to churn" in d
        or "will not upgrade" in d
        or ("tech validation" in d and "but lost" in d)
    ):
        return "churn"
    # Legal blocker — check before no_contact (legal reason can appear in outreach strings)
    if "legal reason" in d or "legal block" in d:
        return "legal_blocker"
    # "Blocked from customer outreach: <reason>" — parse sub-reason first
    if "blocked from customer outreach" in d:
        if (
            "core rep is blocking" in d
            or "technical reason" in d
            or "account team" in d
        ):
            return "core_rep_blocking"
        if "active deal" in d:
            return "active_deal"
        return "no_contact"
    # No contact — cannot reach customer
    if (
        "not able to contact" in d
        or "internal kick-off" in d
        or "no response" in d
        or "refusing to meet" in d
        or "escalating outreach" in d
    ):
        return "no_contact"
    # Core rep / account team blocking
    if (
        "core rep is blocking" in d
        or "account team is blocking" in d
        or "account team decided to delay" in d
        or "ngs sales and core team" in d
    ):
        return "core_rep_blocking"
    # Self-hosted deployment — distinct from generic tech blockers
    if "self-hosted" in d or "self hosted" in d:
        return "self_hosted"
    # Technical blocker
    if (
        "technical reason" in d
        or "technical blocker" in d
        or "tech limitation" in d
        or "behind schedule due to technical" in d
        or "confirmed technical blockers" in d
        or "tenant provision" in d
        or "activation" in d
    ):
        return "tech_blocker"
    # Active deal in flight
    if "active deal" in d:
        return "active_deal"
    # Customer-side delay (non-technical)
    if (
        "pushes to delay" in d
        or "paused by the customer" in d
        or "would like to delay" in d
        or "behind schedule due to customer capacity" in d
        or "customer capacity" in d
        or "customer is not ready" in d
    ):
        return "customer_delay"
    # Legal blocker
    if "legal reason" in d or "legal block" in d:
        return "legal_blocker"
    return ""


def _derive_subtype(status_detail: str, saas_sh: str) -> str:
    """Subtype from status_detail, with SH Only override for blocked accounts."""
    subtype = _subtype_from_detail(status_detail)
    # SH Only accounts blocked by their deployment — override only if not already
    # classified as something stronger (churn takes priority)
    if saas_sh == "SH Only" and subtype not in ("churn", "core_rep_blocking", "legal_blocker", "active_deal"):
        signal = _signal_from_detail(status_detail)
        if signal == "blocked":
            return "self_hosted"
    return subtype


def _status_from_dc(status_detail: str, pc_status: str) -> str:
    """Derive EMEA-style engagement status from DC fields."""
    d = status_detail.lower()
    if "upgrade complete" in d or pc_status == "CC NNL":
        return "Completed"
    if "upgrade started" in d or "upgrade in progress" in d:
        return "In Progress"
    if "tech validation" in d and "won" in d:
        return "Customer Engaged"
    if "customer meeting completed" in d:
        return "Customer Engaged"
    if "customer outreach complete" in d:
        return "Account team contacted"
    if "outreach made" in d:
        return "Account team contacted"
    if pc_status == "Churn":
        return "Churning/Churned"
    if "blocked from internal kick-off" in d:
        return "Account team contacted"
    if "blocked from customer outreach" in d:
        return "Account team contacted"
    return "Account team contacted"


def _clean_cse(val: str) -> str:
    """Return CSE name or empty string. Rejects date-like and junk values."""
    v = val.strip()
    if not v or re.match(r"^\d", v) or "/" in v or "@" in v:
        return ""
    if v.lower() in {"to be hired", "tbd", "n/a", "-"}:
        return ""
    # Known typo in DC file
    if v == "Mikhail Bahkmetiev":
        return "Mikhail Bakhmetiev"
    return v


def _clean_rep(val: str) -> str:
    """Normalise rep/DSM name: strip whitespace and reject placeholder dashes."""
    v = val.strip()
    return "" if v in ("-", "—", "N/A", "n/a", "TBD", "tbd") else v


SUPPORTED_THEATRES = {"EMEA", "JAPAC", "AMER", "LATAM"}


def parse_dc_csv(filepath: Path) -> list[dict]:
    """Parse DC CSE Tracker CSV. Returns records for all supported theatres (EMEA/JAPAC/AMER/LATAM)."""
    records = []
    with open(filepath, newline="", encoding="utf-8-sig", errors="ignore") as f:
        for row in csv.DictReader(f):
            theatre = row.get("account_theatre", "").strip().upper()
            if theatre not in SUPPORTED_THEATRES:
                continue
            aid = row.get("pc_end_customer_account_id", "").strip().lower()
            if not aid:
                continue
            records.append(
                {
                    "account_id": aid,
                    "account_name": row.get("pc_account_name", "").strip(),
                    "account_theatre": theatre,
                    "active_cse": _clean_cse(row.get("CSE Assigned", "")),
                    "dc_assignment": row.get("DC assignment", "").strip(),
                    "owner_e2e": row.get("Owner: End to end upgrade", "").strip(),
                    "dc_status": row.get("PC_CC_Migration_status", "").strip(),
                    "cohort": row.get(
                        "customer_size_cohort_classification", ""
                    ).strip(),
                    "email_sent": row.get("Email sent", "").strip(),
                    "status_detail": row.get("Status Detail", "").strip(),
                    "live_fire": _yn(row.get("Live-fire", "")),
                    # Milestones
                    "m0_complete": _yn(row.get("M0:Internal Kickoff Complete", "")),
                    "m1_complete": _yn(row.get("M1:Customer Outreach Complete", "")),
                    "m1_planned": row.get(
                        "Date - M1:Internal Kickoff Complete", ""
                    ).strip(),
                    "m2_complete": _yn(
                        row.get("M2:Entitlements and Plan aligned with customer", "")
                    ),
                    "m2_planned": row.get(
                        "Date - M2:Entitlements and Plan aligned with customer", ""
                    ).strip(),
                    "m3_complete": _yn(row.get("M3:EB Buy-in Meeting Complete", "")),
                    "m3_planned": row.get("M3 Planned date", "").strip(),
                    "m3_actual": row.get(
                        "Date - M3:EB Buy-in Meeting Complete", ""
                    ).strip(),
                    "m4_complete": _yn(row.get("M4:Discovery complete", "")),
                    "m4_planned": row.get("Date - M4:Discovery complete", "").strip(),
                    "m5_complete": _yn(row.get("M5:Tech validation complete", "")),
                    "m5_planned": row.get(
                        "Date - M5:Tech validation complete", ""
                    ).strip(),
                    "m7_complete": _yn(
                        row.get("M7:Legal and operational upgrade readiness", "")
                    ),
                    "m8_started": _yn(row.get("M8:Upgrade started", "")),
                    "m8_planned": row.get("M8 Planned date", "").strip(),
                    "m8_actual": row.get("Date - M8:Upgrade started", "").strip(),
                    "m9_complete": _yn(row.get("M9:Upgrade complete", "")),
                    "m9_planned": row.get("M9 Planned date", "").strip(),
                    "m9_actual": row.get("Date - M9:Upgrade complete", "").strip(),
                    "upgrade_notes": row.get("Upgrade Notes", "").strip(),
                    "health_notes": row.get("Account Health Notes", "").strip(),
                    "pm_status": row.get("PM Status", "").strip(),
                    "dc_progress": row.get("DC Upgrade Progress Status", "").strip(),
                    "cc_rep": _clean_rep(row.get("cc_Rep (SPO)", "")),
                    "cc_dsm": _clean_rep(row.get("cc_DSM (SPO)", "")),
                    "churn_risk": row.get(
                        "DC Indicated account churn risk", ""
                    ).strip(),
                    # Derived fields — signal, subtype, status from DC Status Detail
                    "signal": _signal_from_detail(row.get("Status Detail", "")),
                    "subtype": _derive_subtype(
                        row.get("Status Detail", ""),
                        row.get("PC_SAAS_vs_SH", "").strip(),
                    ),
                    "status": _status_from_dc(
                        row.get("Status Detail", ""),
                        row.get("PC_CC_Migration_status", ""),
                    ),
                    # Extended fields
                    "last_edited_by": row.get("Last edited by", "").strip(),
                    "last_edited_date": row.get("Last edited date", "").strip(),
                    "roadmap_url": row.get("roadmap", "").strip(),
                    "ps_plan_url": row.get("ps plan", "").strip(),
                    "account_region": row.get("account_region", "").strip(),
                    "current_project_status": row.get(
                        "current_project_status", ""
                    ).strip(),
                    "next_renewal_date": row.get("next_cloud_renewal_date", "").strip(),
                    "past_due_planned": row.get("Past due planned dates", "").strip(),
                    "upgrade_duration_weeks": row.get(
                        "Planned upgrade duration (weeks)", ""
                    ).strip(),
                    "has_partner": row.get("Is there partner", "").strip(),
                    "upgrade_partner": row.get("Upgrade partner name", "").strip(),
                    "m1_details": row.get("M1 Details", "").strip(),
                    "m3_details": row.get("M3 Details", "").strip(),
                    "m5_details": row.get("M5 Details", "").strip(),
                    "milestone_aging": row.get(
                        "Milestone aging calculation", ""
                    ).strip(),
                    "days_since_milestone": row.get(
                        "Days since milestones advanced", ""
                    ).strip(),
                    "momentum_x": row.get("MomentumX", "").strip(),
                    "entitlement_provision": row.get(
                        "entitlement_provision", ""
                    ).strip(),
                    "activation_status": row.get(
                        "activation_tenant_status", ""
                    ).strip(),
                    "posture_workloads": row.get("Posture workloads", "").strip(),
                    "merged_at": datetime.now(timezone.utc).isoformat(),
                }
            )
    logger.info("DC CSE Tracker: parsed %d records across all theatres", len(records))
    return records


def merge_into_state(records: list[dict], state_file: Path) -> dict:
    """
    Merge DC records into state.json by account_id.
    DC CSE Tracker wins on: active_cse, email_sent, all milestone fields.
    Returns summary dict.
    """
    state = json.loads(state_file.read_text(encoding="utf-8"))
    accounts = state.get("accounts", {})

    # Build case-insensitive ID lookup (state keys may be mixed case)
    lower_map = {k.lower(): k for k in accounts}

    matched = 0
    unmatched = []

    for rec in records:
        aid_lower = rec["account_id"]  # already lowercased by parse_dc_csv
        aid = lower_map.get(aid_lower)
        if aid is None:
            unmatched.append(rec["account_name"])
            continue

        acc = accounts[aid]

        # DC always wins on CSE
        if rec["active_cse"]:
            acc["active_cse"] = rec["active_cse"]

        # DC always wins on email_sent (if it has a value)
        if rec["email_sent"]:
            acc["email_sent"] = rec["email_sent"]

        # DC live_fire wins if True
        if rec["live_fire"]:
            acc["live_fire"] = True

        # Store full DC data for milestone access
        acc["dc_data"] = rec
        matched += 1

    state["accounts"] = accounts
    state["dc_last_updated"] = datetime.now(timezone.utc).isoformat()
    state_file.write_text(json.dumps(state, indent=2, ensure_ascii=False))

    logger.info(
        "DC merge: %d/%d matched, %d unmatched", matched, len(records), len(unmatched)
    )
    return {
        "total": len(records),
        "matched": matched,
        "unmatched": len(unmatched),
        "unmatched_list": unmatched,
    }


def load_and_merge(csv_path: Path, state_file: Path) -> dict:
    """One-shot: parse and merge."""
    return merge_into_state(parse_dc_csv(csv_path), state_file)


def parse_unified_xlsx(xlsx_bytes: bytes, conn) -> dict:
    """
    Parse Unified Tracker 2.0 Combined Database tab into accounts + blocked_data.
    Replaces parse_dc_csv as the primary data source.
    """
    import openpyxl
    from io import BytesIO
    from agent.db import wipe_account_data

    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), read_only=True, data_only=True)
    if "Combined Database" not in wb.sheetnames:
        return {"accounts": 0, "changes": 0, "errors": ["Combined Database sheet not found"]}

    ws = wb["Combined Database"]
    rows_iter = ws.iter_rows(min_row=1)
    raw_headers = [str(c.value or "") for c in next(rows_iter)]
    col = {h: i for i, h in enumerate(raw_headers)}

    def _get(row_vals, name, default=""):
        i = col.get(name, -1)
        if i < 0 or i >= len(row_vals):
            return default
        v = row_vals[i]
        return str(v).strip() if v is not None else default

    def _yn(val):
        return 1 if str(val).strip().upper() == "Y" else 0

    wipe_account_data(conn)

    n_accounts = 0
    n_errors = []
    now_iso = datetime.now(timezone.utc).isoformat()

    for row in ws.iter_rows(min_row=2):
        row_vals = [c.value for c in row]
        if not any(v for v in row_vals if v is not None):
            continue

        account_id = _get(row_vals, "pc_end_customer_account_id")
        if not account_id:
            continue

        customer_name = _get(row_vals, "pc_account_name")
        active_cse    = _get(row_vals, "CSE Assigned")
        theatre       = _get(row_vals, "account_theatre")
        sales_region  = _get(row_vals, "account_region")
        cohort        = _get(row_vals, "customer_size_cohort_classification")

        status_detail  = _get(row_vals, "Status Detail")
        subtype        = _subtype_from_detail(status_detail)
        signal         = _signal_from_detail(status_detail)

        # If no signal from emoji, derive from DC progress
        if not signal:
            dc_prog = _get(row_vals, "DC Upgrade Progress Status")
            if dc_prog in ("Yellow", "Red"):
                signal = "at_risk"
            else:
                signal = "green"

        try:
            conn.execute("""
                INSERT OR REPLACE INTO accounts
                  (account_id, customer_name, active_cse, account_theatre, sales_region, created_at)
                VALUES (?,?,?,?,?,?)
            """, (account_id, customer_name, active_cse, theatre, sales_region, now_iso))

            conn.execute("""
                INSERT OR REPLACE INTO blocked_data (
                  account_id, cohort, account_theatre, area, account_region, district,
                  m0_complete, m1_complete, m2_complete, m3_complete, m4_complete,
                  m5_complete, m7_complete, m8_started, m9_complete,
                  m8_actual, m9_actual, m3_planned, m8_planned, m9_planned,
                  upgrade_notes, health_notes, status_detail, dc_progress, churn_risk,
                  cc_rep, cc_dsm, cortexcloud_renewable_acv, pc_cc_migration_status,
                  owner_e2e, dc_assignment, last_edited_by, last_edited_date,
                  current_project_status, field_indicated_churn,
                  signal, subtype
                ) VALUES (
                  ?,?,?,?,?,?,
                  ?,?,?,?,?,
                  ?,?,?,?,
                  ?,?,?,?,?,
                  ?,?,?,?,?,
                  ?,?,?,?,
                  ?,?,?,?,
                  ?,?,
                  ?,?
                )
            """, (
                account_id, cohort, theatre,
                _get(row_vals, "Account_area"),
                _get(row_vals, "account_region"),
                _get(row_vals, "account_district"),
                _yn(_get(row_vals, "M0:Internal Kickoff Complete")),
                _yn(_get(row_vals, "M1:Customer Outreach Complete")),
                _yn(_get(row_vals, "M2:Entitlements and Plan aligned with customer")),
                _yn(_get(row_vals, "M3:EB Buy-in Meeting Complete")),
                _yn(_get(row_vals, "M4:Discovery complete")),
                _yn(_get(row_vals, "M5:Tech validation complete")),
                _yn(_get(row_vals, "M7:Legal and operational upgrade readiness")),
                _yn(_get(row_vals, "M8:Upgrade started")),
                _yn(_get(row_vals, "M9:Upgrade complete")),
                _get(row_vals, "Date - M8:Upgrade started"),
                _get(row_vals, "Date - M9:Upgrade complete"),
                _get(row_vals, "M3 Planned date"),
                _get(row_vals, "M8 Planned date"),
                _get(row_vals, "M9 Planned date"),
                _get(row_vals, "Upgrade Notes"),
                _get(row_vals, "Account Health Notes"),
                status_detail,
                _get(row_vals, "DC Upgrade Progress Status"),
                _get(row_vals, "DC Indicated account churn risk"),
                _get(row_vals, "cc_Rep (SPO)"),   # col 222 (0-indexed)
                _get(row_vals, "DCM"),
                _get(row_vals, "cortexcloud_renewable_acv"),
                _get(row_vals, "PC_CC_Migration_status"),
                _get(row_vals, "Owner: End to end upgrade"),
                _get(row_vals, "DC assignment"),
                _get(row_vals, "Last edited by"),
                _get(row_vals, "Last edited date"),
                _get(row_vals, "current_project_status"),
                _get(row_vals, "Field indicated churn (SPO)"),
                signal, subtype,
            ))
            n_accounts += 1
        except Exception as e:
            n_errors.append(f"{account_id}: {e}")

    wb.close()
    conn.commit()
    return {"accounts": n_accounts, "changes": n_accounts, "errors": n_errors[:10], "error_total": len(n_errors)}
